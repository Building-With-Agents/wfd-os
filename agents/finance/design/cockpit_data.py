"""
CFA Cockpit — Data Extractor

Reads the four K8341 source spreadsheets and produces a single structured
dict that matches the shape expected by the cockpit template.

In production, this module gets replaced by API calls to the grant-compliance
service. The shape of the dict stays the same; only the source changes.
"""

from openpyxl import load_workbook
from datetime import date
from pathlib import Path

import httpx

from agents.finance.audit_dimension_display import display_name_for_role


# Compliance engine endpoint that supplies the six audit-dimension
# rows (metadata + readiness). Server-side call; the Next.js rewrite
# (`/api/grant-compliance/*`) is for browser calls only.
_COMPLIANCE_DIMENSIONS_URL = "http://127.0.0.1:8000/compliance/dimensions"

# Compliance audit activity feed. Fetched once per extract_all()
# cycle; the cockpit's /cockpit/activity endpoint applies label
# translation (agents/finance/audit_activity_labels.py) before
# serving to the UI. See spec §v1.2.9 for the "Recent Compliance
# Activity" panel contract.
_COMPLIANCE_ACTIVITY_URL = "http://127.0.0.1:8000/compliance/activity"

# Per-dimension gap list. Fetched once per extract_all() cycle for
# each of the six canonical dimensions — six sequential GETs. The
# spec documented this endpoint as "lazy" (so it wouldn't bloat the
# dimensions response), not necessarily click-time-lazy; the cockpit
# consumes it at refresh time because the existing drill system
# pre-builds every drill panel in extract_all (see /cockpit/drills/{key}
# which is a straight dict lookup). See spec §v1.2.7 for rationale.
_COMPLIANCE_GAPS_URL_TEMPLATE = (
    "http://127.0.0.1:8000/compliance/dimensions/{dimension_id}/gaps"
)

# Deliberately minimal fallback used only when the compliance engine
# is unreachable. Holds ids and human-readable titles only — every
# other field (auditor-perspective copy, CFR citations, owner role,
# readiness pct) is expected to come from the engine. Keeping this
# short avoids re-hardcoding the dimension contract here.
_FALLBACK_DIMENSION_TITLES: tuple[tuple[str, str], ...] = (
    ("allowable_costs", "Allowable costs"),
    ("transaction_documentation", "Transaction documentation"),
    ("time_effort", "Time & effort certifications"),
    ("procurement", "Procurement & competition"),
    ("subrecipient_monitoring", "Subrecipient monitoring"),
    ("performance_reporting", "Performance reporting accuracy"),
)


def fetch_audit_activity_from_engine(
    days: int = 7, limit: int = 50, timeout_seconds: float = 5.0
) -> dict:
    """Fetch the last `days` of compliance audit log entries from the engine.

    Returns a dict with:
      - entries: list of raw entry dicts as returned by
        GET /compliance/activity (see engine commit e87f3b7).
      - engine_ok: True on success, False on any failure.
      - error: None on success, str(exc) on failure.
      - computed_at: engine's timestamp on success, None on failure.

    Never raises. When engine_ok is False, the cockpit's
    /cockpit/activity endpoint surfaces engine_status="unreachable"
    and the UI shows "Compliance engine unavailable — recent activity
    not available" per spec §v1.2.9.

    TODO(v1.2 cockpit-side step 6 follow-up): add pytest on this
    branch and cover success / timeout / error / empty-entries paths.
    Tests deferred per scope; see integration_notes.md on
    feature/compliance-engine-extract.
    """
    try:
        response = httpx.get(
            _COMPLIANCE_ACTIVITY_URL,
            params={"days": days, "limit": limit},
            timeout=timeout_seconds,
        )
        response.raise_for_status()
        payload = response.json()
        return {
            "entries": payload.get("entries", []),
            "engine_ok": True,
            "error": None,
            "computed_at": payload.get("computed_at"),
        }
    except Exception as exc:  # noqa: BLE001 — intentional catch-all
        print(f"[cockpit] audit activity fetch failed: {exc}")
        return {
            "entries": [],
            "engine_ok": False,
            "error": str(exc),
            "computed_at": None,
        }


def fetch_audit_gaps_from_engine(
    dimension_id: str, timeout_seconds: float = 5.0
) -> dict:
    """Fetch the gap list for one audit dimension from the engine.

    Returns a dict with:
      - gaps: list of gap dicts (engine shape; see spec §v1.2.7).
      - gap_count: integer count.
      - status: "computed" | "placeholder" from the engine.
      - placeholder_message: set when status == "placeholder", else None.
      - engine_ok: True on success, False on any failure.
      - error: None on success, str(exc) on failure.

    Never raises — callers render a degraded "Engine unreachable"
    drill-panel section when engine_ok is False.

    TODO(v1.2 cockpit-side step 4 follow-up): add pytest on this
    branch and cover success / timeout / placeholder / computed paths.
    Tests deferred per scope; see integration_notes.md on
    feature/compliance-engine-extract.
    """
    url = _COMPLIANCE_GAPS_URL_TEMPLATE.format(dimension_id=dimension_id)
    try:
        response = httpx.get(url, timeout=timeout_seconds)
        response.raise_for_status()
        payload = response.json()
        return {
            "gaps": payload.get("gaps", []),
            "gap_count": payload.get("gap_count", 0),
            "status": payload.get("status", "placeholder"),
            "placeholder_message": payload.get("placeholder_message"),
            "engine_ok": True,
            "error": None,
        }
    except Exception as exc:  # noqa: BLE001 — intentional catch-all
        print(f"[cockpit] audit gap fetch failed for {dimension_id!r}: {exc}")
        return {
            "gaps": [],
            "gap_count": 0,
            "status": "placeholder",
            "placeholder_message": None,
            "engine_ok": False,
            "error": str(exc),
        }


def _fmt_gap_dollars(amount_dollars) -> str:
    """Format a dollar amount for drill-panel gap labels.

    Mirrors the engine's own formatter (both rendered as "$X,XXX" for
    whole numbers, "$X,XXX.YY" otherwise) so amounts inside
    compliance-flag gap metadata display the same way the gap title
    already does for transaction_documentation gaps.

    TODO: share a single formatter between engine and cockpit when
    wfdos-common lands; for now the helper is duplicated across the
    two branches by design. Cockpit-side TS uses fmtUSD for the
    DocumentationGap stat card — separate surface, same "$X,XXX"
    output shape.
    """
    if not isinstance(amount_dollars, (int, float)):
        return "—"
    if amount_dollars == int(amount_dollars):
        return f"${int(amount_dollars):,}"
    return f"${amount_dollars:,.2f}"


def _build_audit_gap_section(gap_response: dict) -> dict:
    """Build the drill-panel section that replaces the old "Open gaps"
    placeholder. Three branches: engine-unreachable (prose with
    operational guidance), placeholder dimension (prose with the
    engine's honest placeholder_message), computed dimension (rows
    with one row per gap, or a single "No open gaps" row when clean).
    """
    if not gap_response.get("engine_ok"):
        return {
            "id": "open_gaps",
            "type": "prose",
            "title": "Open gaps",
            "body": (
                "Engine unreachable — gap detail is not available. "
                "Verify the compliance engine is running."
            ),
        }

    status = gap_response.get("status", "placeholder")
    if status == "placeholder":
        message = gap_response.get("placeholder_message") or (
            "Gap detection not yet available for this dimension."
        )
        return {
            "id": "open_gaps",
            "type": "prose",
            "title": "Open gaps",
            "body": message,
        }

    # Computed dimension.
    gaps = gap_response.get("gaps") or []
    if not gaps:
        return {
            "id": "open_gaps",
            "type": "rows",
            "title": "Open gaps",
            "rows": [{"label": "No open gaps in this dimension.", "value": "✓"}],
        }

    rows: list[dict] = []
    for gap in gaps:
        meta = gap.get("metadata") or {}
        gap_type = gap.get("type")
        title = gap.get("title", "")
        detail = gap.get("detail", "")
        if gap_type == "compliance_flag":
            vendor = meta.get("vendor_name") or "—"
            amt_str = _fmt_gap_dollars(meta.get("amount_dollars"))
            date_str = meta.get("txn_date") or "—"
            label = (
                f"{title} — {detail} "
                f"({vendor} · {amt_str} · {date_str})"
            )
            value = (meta.get("severity") or "").upper()
        elif gap_type == "missing_documentation":
            # Engine pre-formats title as "{vendor} — $X,XXX"; append
            # date + qb_id for cross-reference at the end of the label.
            date_str = meta.get("txn_date") or "—"
            qb_id = meta.get("qb_id") or ""
            suffix = f" ({qb_id})" if qb_id else ""
            label = f"{title} · {date_str}{suffix}"
            value = detail
        else:
            # Unknown type — render generically rather than hide.
            label = title
            value = detail
        rows.append({"label": label, "value": value})

    return {
        "id": "open_gaps",
        "type": "rows",
        "title": "Open gaps",
        "rows": rows,
    }


def _unreachable_stats_fallback() -> dict:
    """Stats payload the cockpit renders when the compliance engine is
    unreachable. Shape mirrors the engine's `stats` block. Per spec
    §v1.2.6, `te_certs_status` is "engine_unreachable" (distinct from
    the normal "placeholder_pending_employee_grant_data") so the UI
    can distinguish "not measured yet" (roadmap) from "engine is down"
    (operational)."""
    return {
        "overall_readiness_pct": None,
        "overall_readiness_basis": {
            "computed_dimension_count": 0,
            "total_dimension_count": 6,
        },
        "doc_gap_count": None,
        # 250_000 cents = $2,500 de minimis threshold. Documented
        # default; mirrored here so the cockpit can still render the
        # "Transactions over $2,500" label even when the engine is down.
        "doc_gap_threshold_cents": 250_000,
        "te_certs_status": "engine_unreachable",
    }


def fetch_audit_dimensions_from_engine(timeout_seconds: float = 5.0) -> dict:
    """Fetch the six audit dimensions and stat-card values from the engine.

    Returns a dict with:
      - dimensions: list of dimension dicts (engine shape), or the
        minimal id+title fallback list when the engine is unreachable.
      - stats: dict matching the engine's stats block shape. When the
        engine is unreachable or its response omits the field, returns
        the _unreachable_stats_fallback() payload with
        te_certs_status="engine_unreachable". See spec §v1.2.6.
      - engine_ok: True on success, False on any failure.
      - error: None on success, str(exc) on failure.
      - computed_at: engine's timestamp on success, None on failure.

    Never raises — all exceptions are caught and surfaced via
    engine_ok=False. Callers render a clearly-degraded state rather
    than pretending everything is fine.

    TODO(v1.2 cockpit-side step 2/3 follow-up): add pytest on this
    branch and cover the success / timeout / error / missing-stats
    paths. Tests deferred per integration_notes.md on
    feature/compliance-engine-extract.
    """
    try:
        response = httpx.get(_COMPLIANCE_DIMENSIONS_URL, timeout=timeout_seconds)
        response.raise_for_status()
        payload = response.json()
        # Engine responded; trust its stats block when present. If the
        # engine is an older version without stats, fall back to the
        # same unreachable-shaped payload — not perfectly accurate
        # (engine IS technically reachable) but a tolerable edge case
        # given the current engine always emits stats (see commit 643d9d8).
        stats = payload.get("stats") or _unreachable_stats_fallback()
        return {
            "dimensions": payload.get("dimensions", []),
            "stats": stats,
            "engine_ok": True,
            "error": None,
            "computed_at": payload.get("computed_at"),
        }
    except Exception as exc:  # noqa: BLE001 — intentional catch-all for degraded rendering
        fallback = [
            {
                "id": did,
                "title": title,
                "what_auditors_look_for":
                    "Engine unreachable — readiness data not available.",
                "cfr_citations": [],
                "compliance_supplement_area": "",
                "owner_role": "",
                "default_tone": "neutral",
                "readiness_pct": None,
                "status": "placeholder",
            }
            for did, title in _FALLBACK_DIMENSION_TITLES
        ]
        print(f"[cockpit] audit dimensions fetch failed: {exc}")
        return {
            "dimensions": fallback,
            "stats": _unreachable_stats_fallback(),
            "engine_ok": False,
            "error": str(exc),
            "computed_at": None,
        }


# =============================================================================
# Drill content schema — polymorphic section types
# =============================================================================
# Every entry in the drillRegistry conforms to this shape. Section types are
# discriminated by the "type" field: `rows`, `table`, `chart`, `prose`,
# `verdict`, `timeline`, `action_items`. Each type has its own renderer in
# cockpit_template.html. Tone values map to CSS custom properties (--good,
# --watch, --critical, --text-3 for neutral) so every color decision lives
# in data, not in template hex ternaries.
#
# Validation runs at the end of build_drills() and fails loudly on any
# nonconformant entry — catching schema drift at build time rather than
# silently rendering a broken drill in the browser.

TONE_VALUES = ("good", "watch", "critical", "neutral")
SECTION_TYPES = (
    "rows", "table", "chart", "prose",
    "verdict", "timeline", "action_items",
)


def _require(cond: bool, ctx: str, msg: str) -> None:
    if not cond:
        raise ValueError(f"Drill schema violation at {ctx}: {msg}")


def _validate_tone(tone, ctx: str) -> None:
    _require(tone in TONE_VALUES, ctx,
             f"tone must be one of {TONE_VALUES}, got {tone!r}")


def _validate_rows(section: dict, ctx: str) -> None:
    _require(isinstance(section.get("title"), str), ctx, "title required")
    rows = section.get("rows", [])
    _require(isinstance(rows, list), ctx, "rows must be a list")
    for i, row in enumerate(rows):
        rctx = f"{ctx}.rows[{i}]"
        _require(isinstance(row, dict), rctx, "row must be a dict")
        _require("label" in row and "value" in row, rctx,
                 "row must have label + value")


def _validate_table(section: dict, ctx: str) -> None:
    _require(isinstance(section.get("title"), str), ctx, "title required")
    cols = section.get("columns", [])
    _require(isinstance(cols, list) and cols, ctx, "columns must be non-empty")
    for i, c in enumerate(cols):
        cctx = f"{ctx}.columns[{i}]"
        _require(isinstance(c, dict), cctx, "column must be a dict")
        _require("key" in c and "label" in c, cctx, "column needs key + label")
    rows = section.get("rows", [])
    _require(isinstance(rows, list), ctx, "rows must be a list")


def _validate_chart(section: dict, ctx: str) -> None:
    _require(isinstance(section.get("title"), str), ctx, "title required")
    _require(section.get("chart_type") in ("bar", "line", "area"), ctx,
             "chart_type must be one of bar|line|area")
    for key in ("x_axis", "y_axis"):
        ax = section.get(key, {})
        _require(isinstance(ax, dict) and "key" in ax, f"{ctx}.{key}",
                 "axis must be dict with at least a 'key' field")
    _require(isinstance(section.get("data"), list), ctx,
             "data must be a list")
    for i, ref in enumerate(section.get("reference_lines") or []):
        rctx = f"{ctx}.reference_lines[{i}]"
        _require(isinstance(ref.get("value"), (int, float)), rctx,
                 "reference line needs numeric value")
        if "tone" in ref:
            _validate_tone(ref["tone"], rctx)


def _validate_prose(section: dict, ctx: str) -> None:
    _require(isinstance(section.get("body"), str) and section["body"], ctx,
             "prose.body must be a non-empty string")


def _validate_verdict(section: dict, ctx: str) -> None:
    _validate_tone(section.get("tone"), ctx)
    _require(isinstance(section.get("headline"), str) and section["headline"],
             ctx, "verdict.headline required")
    _require(isinstance(section.get("body"), str) and section["body"],
             ctx, "verdict.body required")


def _validate_timeline(section: dict, ctx: str) -> None:
    _require(isinstance(section.get("title"), str), ctx, "title required")
    events = section.get("events", [])
    _require(isinstance(events, list), ctx, "events must be a list")
    for i, e in enumerate(events):
        ectx = f"{ctx}.events[{i}]"
        _require(isinstance(e, dict), ectx, "event must be a dict")
        _require("date" in e and "title" in e, ectx,
                 "event needs date + title")
        if "tone" in e:
            _validate_tone(e["tone"], ectx)


def _validate_action_items(section: dict, ctx: str) -> None:
    _require(isinstance(section.get("title"), str), ctx, "title required")
    items = section.get("items", [])
    _require(isinstance(items, list), ctx, "items must be a list")
    for i, it in enumerate(items):
        ictx = f"{ctx}.items[{i}]"
        _require(isinstance(it, dict), ictx, "item must be a dict")
        _require(it.get("priority") in ("HIGH", "MEDIUM", "LOW"), ictx,
                 "item.priority must be HIGH|MEDIUM|LOW")
        _require("text" in it, ictx, "item needs text")


_SECTION_VALIDATORS = {
    "rows": _validate_rows,
    "table": _validate_table,
    "chart": _validate_chart,
    "prose": _validate_prose,
    "verdict": _validate_verdict,
    "timeline": _validate_timeline,
    "action_items": _validate_action_items,
}


def validate_drill(key: str, entry: dict) -> None:
    """Validate a single drill registry entry. Raises ValueError on mismatch.

    Enforces the shape documented above — required top-level fields, known
    section types, and per-type constraints. Optional fields (status_chip,
    actions, updated_at, source) are validated shallowly when present.
    """
    ctx = f"drill {key!r}"
    _require(isinstance(entry, dict), ctx, "entry must be a dict")
    for f in ("eyebrow", "title", "summary"):
        _require(isinstance(entry.get(f), str), ctx, f"{f} must be a string")
    sections = entry.get("sections")
    _require(isinstance(sections, list), ctx, "sections must be a list")
    # Section ids must be present, non-empty strings, and unique within a
    # drill. The drill-chat surface cites them as `sources: ["<id>", ...]`
    # in its structured response; the frontend scrolls to the matching
    # section. See agents/finance/design/chat_spec.md §"Section IDs".
    seen_ids: set[str] = set()
    for i, section in enumerate(sections):
        sctx = f"{ctx}.sections[{i}]"
        _require(isinstance(section, dict), sctx, "section must be a dict")
        sid = section.get("id")
        _require(isinstance(sid, str) and sid, sctx,
                 "id must be a non-empty string")
        _require(sid not in seen_ids, sctx,
                 f"duplicate section id within drill: {sid!r}")
        seen_ids.add(sid)
        stype = section.get("type")
        _require(stype in SECTION_TYPES, sctx,
                 f"type must be one of {SECTION_TYPES}, got {stype!r}")
        _SECTION_VALIDATORS[stype](section, sctx)
    chip = entry.get("status_chip")
    if chip is not None:
        _require(isinstance(chip, dict), ctx, "status_chip must be a dict")
        _require(isinstance(chip.get("label"), str), ctx,
                 "status_chip.label required")
        _validate_tone(chip.get("tone"), f"{ctx}.status_chip")
    actions = entry.get("actions")
    if actions is not None:
        _require(isinstance(actions, list), ctx, "actions must be a list")
        for i, a in enumerate(actions):
            actx = f"{ctx}.actions[{i}]"
            _require(isinstance(a.get("label"), str), actx, "label required")
            _require(a.get("intent") in ("navigate", "chat", "export"), actx,
                     "intent must be navigate|chat|export")


def validate_registry(registry: dict) -> None:
    """Validate every entry in the drill registry. Fails loudly on first
    nonconformant entry, collecting errors into a single message."""
    errors = []
    for key, entry in registry.items():
        try:
            validate_drill(key, entry)
        except ValueError as e:
            errors.append(str(e))
    if errors:
        raise ValueError(
            "Drill registry failed validation:\n  - "
            + "\n  - ".join(errors)
        )


# Source spreadsheets and downstream surfaces (Transactions, Q1 reimbursement,
# Provider Reconciliation) use longer provider names than the canonical ones
# build_drills() emits into the drill registry. Canonicalize on every
# data-drill emission so clicks always resolve to the registered drill entry.
PROVIDER_CANONICAL = {
    "Vets2Tech / St. Martin University": "Vets2Tech",
    "St Martins - Washington Vets 2 Tech": "Vets2Tech",
    "Year Up Puget Sound": "Year Up",
    "Code Day X Mint": "Code Day",
    "Code Day X MinT": "Code Day",
    "Code Day / MinT": "Code Day",
    "CodeDay/MinT": "Code Day",
    "PNW Cyber Challenge": "PNW CCG",
    "NCESD 171": "NCESD",
    "Riipen / North Seattle College": "Riipen",
    "Ada Developers": "Ada",
    "Ada Developers Academy": "Ada",
    "AI Engage Group LLC": "AI Engage",
    "Pete & Kelly Vargo": "CFA Contractors (Pete & Kelly Vargo)",
}


def canonical_provider(name: str) -> str:
    """Return the canonical drill-registry key for a provider name."""
    if name is None:
        return ""
    return PROVIDER_CANONICAL.get(name, name)


def extract_providers(recon_path: Path) -> dict:
    """Read v3 Provider Reconciliation. Returns providers grouped by status."""
    wb = load_workbook(recon_path, data_only=True)
    ws = wb["Provider Reconciliation"]

    # Exact group header text → group key
    group_markers = {
        "ESD-DIRECTED CLAWBACKS": "terminated",
        "ACTIVE TRAINING PROVIDERS": "active",
        "CLOSED PROVIDERS — DELIVERED PLACEMENTS": "closed_with_placements",
        "CLOSED PROVIDERS — SUPPORT / ENGAGEMENT": "closed_support",
        "CFA CONTRACTORS — incl AI Engage": "cfa_contractors",
    }
    # When we hit one of these, stop assigning rows to any group
    stop_markers = {
        "CFA SALARIES & BENEFITS",
        "ITEMS REQUIRING IMMEDIATE CLARIFICATION",
        "FULL GRANT TOTAL",
    }
    # Subtotal / header rows to skip even when current_group is set
    skip_rows = {
        "Provider", "CLAWBACK SUBTOTAL", "ACTIVE PROVIDERS SUBTOTAL",
        "CLOSED WITH PLACEMENTS SUBTOTAL", "SUPPORT PROVIDERS SUBTOTAL",
        "CFA CONTRACTORS SUBTOTAL",
    }

    # Operational categorization (orthogonal to status grouping):
    # Training providers run programs that produce placements; strategic
    # partners do engagement, recruitment, and convening work that supports
    # the coalition but doesn't directly place candidates.
    STRATEGIC_PARTNERS = {
        "WABS", "Evergreen Goodwill", "Seattle Jobs Initiative",
        "WTIA", "ESD 112", "I&CT (Bellevue College)", "DynaTech Systems",
    }

    providers = {k: [] for k in group_markers.values()}
    current_group = None

    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        cell0 = str(row[0]).strip()

        # Stop markers exit data extraction entirely
        if any(cell0.startswith(m) for m in stop_markers):
            current_group = None
            continue

        # Group headers update current group (require exact prefix match)
        matched_group = next(
            (k for marker, k in group_markers.items() if cell0.startswith(marker)),
            None,
        )
        if matched_group:
            current_group = matched_group
            continue

        # Skip subtotal/header rows
        if cell0 in skip_rows:
            continue

        # Add as data row if we're in a group
        if current_group:
            # Classify operationally
            if current_group == "cfa_contractors":
                category = "cfa_contractor"
            elif any(cell0.startswith(s) for s in STRATEGIC_PARTNERS):
                category = "strategic"
            else:
                category = "training"

            providers[current_group].append({
                "name": cell0,
                "budget": _to_float(row[1]),
                "qb_actual": _to_float(row[3]),
                "balance": _to_float(row[4]),
                "notes": str(row[5])[:120] if row[5] else "",
                "category": category,
            })

    wb.close()
    return providers


def extract_action_items(recon_path: Path) -> list:
    """Read the Action Items sheet from v3 reconciliation."""
    wb = load_workbook(recon_path, data_only=True)
    ws = wb["Action Items"]
    items = []
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        cell0 = str(row[0]).strip().upper()
        if cell0 in ("HIGH", "MEDIUM", "LOW"):
            items.append({
                "priority": cell0,
                "area": str(row[1]) if row[1] else "",
                "action": str(row[2])[:200] if row[2] else "",
                "owner": str(row[3]) if row[3] else "",
            })
    wb.close()
    return items


def extract_cost_per_placement(cpp_path: Path) -> dict:
    """Read Cost Per Placement sheet — quarterly payments + CPP per provider."""
    wb = load_workbook(cpp_path, data_only=True)
    ws = wb["Cost Per Placement"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    providers = []
    for row in rows[5:13]:  # data rows 6-13 in 1-indexed (Ada through NCESD)
        if row[0] is None:
            continue
        providers.append({
            "name": str(row[0]),
            "quarterly": [_to_float(row[i]) for i in range(1, 9)],
            "total_paid": _to_float(row[9]),
            "q1_26_retraction": _to_float(row[10]),
            "total_placements": int(row[11]) if row[11] else 0,
            "net_placements": int(row[12]) if row[12] else 0,
            "cpp": _to_float(row[13]),
        })

    # totals row (R14)
    total_row = rows[13]
    totals = {
        "total_paid": _to_float(total_row[9]),
        "total_placements": int(total_row[11]) if total_row[11] else 0,
        "net_placements": int(total_row[12]) if total_row[12] else 0,
        "weighted_cpp": _to_float(total_row[13]),
    }

    return {"providers": providers, "totals": totals}


def extract_budget_categories(exhb_path: Path) -> dict:
    """Read Exhibit B Amendment 1 — line item budget categories."""
    wb = load_workbook(exhb_path, data_only=True)
    ws = wb["Line Item Detail"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # rows 12-19 are the 8 line items (1-indexed)
    line_items = {}
    for r in rows[11:19]:
        if r[1]:
            line_items[str(r[1]).strip()] = _to_float(r[6])

    return {
        "salaries": line_items.get("Personnel: Salaries", 0),
        "benefits": line_items.get("Personnel: Benefits", 0),
        "travel": line_items.get("Travel", 0),
        "communications": line_items.get("Communications", 0),
        "supplies": line_items.get("Supplies", 0),
        "other": line_items.get("Other", 0),
        "subcontracts": line_items.get("Subcontracts", 0),
        "indirect": line_items.get("Indirect Costs", 0),
        "total": _to_float(rows[19][6]) if len(rows) > 19 else 4_875_000,
    }


def compute_summary(providers: dict, cpp: dict, budget: dict) -> dict:
    """Derive headline numbers from extracted source data."""
    # Total paid to GJC contractors (training + support providers)
    all_gjc = [
        p for group in ["terminated", "active",
                        "closed_with_placements", "closed_support"]
        for p in providers[group]
    ]
    gjc_paid = sum(p["qb_actual"] for p in all_gjc)
    gjc_budget = sum(p["budget"] for p in all_gjc)

    # Split GJC into Training vs. Strategic
    training_providers = [p for p in all_gjc if p["category"] == "training"]
    strategic_providers = [p for p in all_gjc if p["category"] == "strategic"]
    training_budget = sum(p["budget"] for p in training_providers)
    training_paid = sum(p["qb_actual"] for p in training_providers)
    strategic_budget = sum(p["budget"] for p in strategic_providers)
    strategic_paid = sum(p["qb_actual"] for p in strategic_providers)

    cfa_contractor_paid = sum(p["qb_actual"] for p in providers["cfa_contractors"])
    cfa_contractor_budget = sum(p["budget"] for p in providers["cfa_contractors"])

    # Backbone subtotals — staff & overhead (4 categories from Exh B)
    other_direct_total = (
        budget["travel"] + budget["communications"]
        + budget["supplies"] + budget["other"]
    )
    backbone_budget = (
        budget["salaries"] + budget["benefits"]
        + other_direct_total + budget["indirect"]
    )
    # From v3 reconciliation: salaries/overhead subtotal QB total $1,227,850.82
    # Balance remaining = $310,701.50
    backbone_qb_total = 1_227_850.82
    backbone_remaining = backbone_budget - backbone_qb_total

    # Months remaining to grant end (Sept 30, 2026)
    today = date.today()
    grant_end = date(2026, 9, 30)
    days_remaining = (grant_end - today).days
    months_remaining = round(days_remaining / 30.4, 1)

    return {
        "today": today.strftime("%B %d, %Y"),
        "days_remaining": days_remaining,
        "months_remaining": months_remaining,
        "grant_total_budget": budget["total"],
        "gjc_budget": gjc_budget,
        "gjc_paid": gjc_paid,
        "gjc_remaining": gjc_budget - gjc_paid,
        "training_budget": training_budget,
        "training_paid": training_paid,
        "training_remaining": training_budget - training_paid,
        "strategic_budget": strategic_budget,
        "strategic_paid": strategic_paid,
        "strategic_remaining": strategic_budget - strategic_paid,
        "cfa_contractor_budget": cfa_contractor_budget,
        "cfa_contractor_paid": cfa_contractor_paid,
        "cfa_contractor_remaining": cfa_contractor_budget - cfa_contractor_paid,
        "backbone_budget": backbone_budget,
        "backbone_qb_paid": backbone_qb_total,
        "backbone_remaining": backbone_remaining,
        "backbone_runway_combined": backbone_remaining + (cfa_contractor_budget - cfa_contractor_paid),
        # category-level (pro-rated since QB only gives lump sum)
        "categories": [
            {"name": "GJC Contractors — Training Providers",
             "budget": training_budget, "spent": training_paid,
             "remaining": training_budget - training_paid,
             "pct": training_paid / training_budget * 100 if training_budget else 0,
             "color": "green",
             "note": f"{len(training_providers)} providers · placement-based"},
            {"name": "GJC Contractors — Strategic Partners",
             "budget": strategic_budget, "spent": strategic_paid,
             "remaining": strategic_budget - strategic_paid,
             "pct": strategic_paid / strategic_budget * 100 if strategic_budget else 0,
             "color": "green",
             "note": f"{len(strategic_providers)} partners · engagement, recruitment, convening"},
            {"name": "CFA Contractors · AI Engage + P&K",
             "budget": cfa_contractor_budget, "spent": cfa_contractor_paid,
             "remaining": cfa_contractor_budget - cfa_contractor_paid,
             "pct": cfa_contractor_paid / cfa_contractor_budget * 100 if cfa_contractor_budget else 0,
             "color": "amber"},
            {"name": "Personnel — Salaries",
             "budget": budget["salaries"],
             "spent": budget["salaries"] * (backbone_qb_total / backbone_budget),
             "remaining": budget["salaries"] - (budget["salaries"] * (backbone_qb_total / backbone_budget)),
             "pct": backbone_qb_total / backbone_budget * 100,
             "color": "amber",
             "prorated": True},
            {"name": "Personnel — Benefits",
             "budget": budget["benefits"],
             "spent": budget["benefits"] * (backbone_qb_total / backbone_budget),
             "remaining": budget["benefits"] - (budget["benefits"] * (backbone_qb_total / backbone_budget)),
             "pct": backbone_qb_total / backbone_budget * 100,
             "color": "amber",
             "prorated": True},
            {"name": "Other Direct Costs",
             "budget": other_direct_total,
             "spent": other_direct_total * (backbone_qb_total / backbone_budget),
             "remaining": other_direct_total - (other_direct_total * (backbone_qb_total / backbone_budget)),
             "pct": backbone_qb_total / backbone_budget * 100,
             "color": "amber",
             "prorated": True},
            {"name": "Indirect Costs",
             "budget": budget["indirect"],
             "spent": budget["indirect"] * (backbone_qb_total / backbone_budget),
             "remaining": budget["indirect"] - (budget["indirect"] * (backbone_qb_total / backbone_budget)),
             "pct": backbone_qb_total / backbone_budget * 100,
             "color": "amber",
             "prorated": True},
        ],
    }


def _cpp_tone(true_cpp: float) -> str:
    """Map a true cost-per-placement to a tone band. Single source of
    truth for CPP color decisions — no Jinja hex ternaries."""
    if true_cpp <= 0:
        return "neutral"
    if true_cpp <= 2500:
        return "good"
    if true_cpp <= 4000:
        return "watch"
    return "critical"


def _cpp_band_label(true_cpp: float) -> str:
    """Human-readable label for the CPP threshold band."""
    if true_cpp <= 0:
        return "No CPP — no payments recorded"
    if true_cpp <= 2500:
        return "Green (highly cost-effective)"
    if true_cpp <= 4000:
        return "Amber (acceptable)"
    return "Red (problem)"


def _placements_per_quarter(fp: dict, quarter_labels: list) -> list:
    """Flatten quarterly placements + Q1'26 actuals into a list of
    {period, placements, paid} dicts suitable for the chart/table sections."""
    rows = []
    for i, label in enumerate(quarter_labels):
        rows.append({
            "period": label,
            "placements": fp["quarterly_placements"][i],
            "paid": fp["quarterly_payments"][i],
        })
    # Append Q1'26 actual. Retraction folded into actual for chart clarity.
    q1_26_net = fp["q1_26_placements"] + fp["q1_26_retraction"]
    if q1_26_net or fp.get("q1_26_invoice"):
        rows.append({
            "period": "Q1'26",
            "placements": q1_26_net,
            "paid": fp.get("q1_26_invoice") or 0,
        })
    return rows


def _placements_table_section(fp: dict, quarter_labels: list) -> dict:
    """Per-quarter placements + payments table for a provider drill."""
    quarterly = _placements_per_quarter(fp, quarter_labels)
    return {
        "id": "placements_table",
        "type": "table",
        "title": "Placements by quarter",
        "columns": [
            {"key": "period", "label": "Period"},
            {"key": "placements", "label": "Placements", "align": "right", "numeric": True},
            {"key": "paid", "label": "Paid", "align": "right", "numeric": True},
        ],
        "rows": [
            {"period": r["period"],
             "placements": r["placements"],
             "paid": f"${r['paid']:,.0f}" if r["paid"] else "—"}
            for r in quarterly
        ] + [{
            "period": "Total",
            "placements": fp["total_placements_net"],
            "paid": f"${fp['total_paid']:,.0f}",
            "total": True,
        }],
    }


def _placements_chart_section(fp: dict, quarter_labels: list,
                              target) -> dict:
    """Bar chart of placements over time, with an optional target
    reference line tone-keyed to how close the provider got.

    Negative values (retractions) are dropped from the chart — they
    belong in the table where they have column context. The chart
    represents positive delivery by quarter.
    """
    quarterly = [r for r in _placements_per_quarter(fp, quarter_labels)
                 if r["placements"] >= 0]
    share = (target / len(quarterly)) if (target and quarterly) else 0
    data = [{
        "period": r["period"],
        "placements": r["placements"],
        "tone": "good" if share and r["placements"] >= share
                else ("watch" if r["placements"] > 0 else "neutral"),
    } for r in quarterly]
    section = {
        "id": "placements_chart",
        "type": "chart",
        "title": "Placements over time",
        "chart_type": "bar",
        "x_axis": {"key": "period", "label": "Quarter"},
        "y_axis": {"key": "placements", "label": "Count"},
        "data": data,
    }
    if target and target > 0:
        reached = fp["total_placements_net"] + fp.get("recovered", 0)
        tone = "good" if reached >= target else ("watch" if reached >= target * 0.75 else "critical")
        section["reference_lines"] = [
            {"value": target, "label": "Target", "tone": tone}
        ]
    return section


def _cost_analysis_rows_section(fp: dict) -> dict:
    """Rows-type cost analysis — for providers in the green band."""
    return {
        "id": "cost_analysis",
        "type": "rows",
        "title": "Cost Analysis",
        "rows": [
            {"label": "Reported-only CPP", "value": f"${fp['cpp']:,.0f}"},
            {"label": "True CPP (incl. recovery)",
             "value": f"${fp['true_cpp']:,.0f}", "emphasize": True},
            {"label": "Threshold band", "value": _cpp_band_label(fp["true_cpp"])},
        ],
    }


def _cost_analysis_verdict_section(fp: dict) -> dict:
    """Verdict-type cost analysis — for providers in the amber/red bands.

    Voice: data-forward. State the number, state the threshold relation,
    let the reader draw the conclusion. No editorializing verbs like
    "exceeds" or "sits between" — just the facts.
    """
    tone = _cpp_tone(fp["true_cpp"])
    pieces = [f"{fp['total_placements_net']} placements reported"]
    if fp.get("recovered"):
        pieces.append(f"plus {fp['recovered']} recovered through CFA outreach")
    placements_clause = ", ".join(pieces)
    if fp.get("recovered") and fp["true_cpp"] != fp["cpp"]:
        cpp_clause = (
            f"Reported-only CPP was ${fp['cpp']:,.0f}; including recovery, "
            f"true CPP is ${fp['true_cpp']:,.0f}."
        )
    else:
        cpp_clause = f"True CPP: ${fp['true_cpp']:,.0f}."
    body = (
        f"{placements_clause} on ${fp['total_paid']:,.0f} paid. {cpp_clause}"
    )
    if tone == "critical":
        headline = f"True CPP ${fp['true_cpp']:,.0f} — in the red band (>$4k)."
    else:  # watch
        headline = f"True CPP ${fp['true_cpp']:,.0f} — in the amber band ($2.5k–$4k)."
    return {
        "id": "cost_analysis",
        "type": "verdict",
        "tone": tone,
        "headline": headline,
        "body": body,
    }


def _action_items_section_for(related: list, title: str | None = None) -> dict:
    """Action-items-type section filtered to items that mention the entity."""
    return {
        "id": "action_items",
        "type": "action_items",
        "title": title or f"Open action items ({len(related)})",
        "items": [{
            "priority": item["priority"],
            "owner": item["owner"],
            "text": item["action"],
        } for item in related],
    }


def _contract_and_spend_section(budget_row: dict) -> dict:
    """Rows section for provider Contract & Spend — same layout as before
    but extracted for reuse across the migrated providers."""
    return {
        "id": "contract_and_spend",
        "type": "rows",
        "title": "Contract & Spend",
        "rows": [
            {"label": "Amended budget (Exh B Amend 1)", "value": f"${budget_row['budget']:,.0f}"},
            {"label": "Paid to date (QB 3/26/26)", "value": f"${budget_row['qb_actual']:,.0f}"},
            {"label": "Balance remaining", "value": f"${budget_row['balance']:,.0f}",
             "emphasize": budget_row['balance'] != 0},
            {"label": "Status group", "value": budget_row['group'].replace('_', ' ').title()},
        ],
        "note": budget_row.get("notes", ""),
    }


def build_drills(data: dict) -> dict:
    """Build drill-down content for every entity worth exploring.

    Keys follow the convention 'type:id' — e.g. 'provider:Ada', 'category:Training',
    'quarter:Q1-26', 'decision:0'. Values are dicts with eyebrow/title/summary/sections
    that the template renders into the slide-out drill panel.
    """
    drills = {}
    fp_by_provider = {r["provider"]: r for r in data["financial_performance"]}
    all_providers = {}
    for group, plist in data["providers"].items():
        for p in plist:
            all_providers[p["name"]] = {**p, "group": group}
    # financial_performance uses canonical names ("NCESD"), while
    # all_providers is keyed by the raw reconciliation name ("NCESD 171").
    # Map canonical → budget row so per-provider drills can find Contract &
    # Spend data even when the name was canonicalized.
    budget_by_canonical = {
        canonical_provider(name): row
        for name, row in all_providers.items()
    }

    # ---------- Per-provider drills ----------
    # Every provider in financial_performance goes through the polymorphic
    # path: Contract & Spend (rows) + Placements (table) + Placements chart
    # + Cost Analysis (rows if green, verdict if amber/red) + related
    # Action Items.
    quarter_labels = data["placements"]["quarter_labels"]
    target_by_provider = {
        r["provider"]: r["target"]
        for r in data["placements"]["quarterly_placements"]
    }
    for fp in data["financial_performance"]:
        name = fp["provider"]
        # Resolve the budget row both by raw name (direct match) and by
        # canonical name (for renamed providers like "NCESD" ← "NCESD 171").
        budget_row = all_providers.get(name) or budget_by_canonical.get(name)
        related_items = [
            item for item in data["action_items"]
            if name.lower() in item["area"].lower() or name.lower() in item["action"].lower()
        ]
        sections = []
        if budget_row:
            sections.append(_contract_and_spend_section(budget_row))
        elif name == "Recovery Operation (AIE + P&K)":
            sections.append({
                "id": "contract_and_spend",
                "type": "rows",
                "title": "Recovery Operation — AI Engage + Pete & Kelly Vargo",
                "rows": [
                    {"label": "AI Engage budget", "value": "$245,000"},
                    {"label": "P&K Vargo budget", "value": "$775,823"},
                    {"label": "Combined paid to date", "value": f"${fp['total_paid']:,.0f}"},
                    {"label": "Placements recovered", "value": str(fp["total_placements_net"])},
                    {"label": "Cost per recovery", "value": f"${fp['cpp']:,.0f}"},
                ],
                "note": "Joint recovery operation — AIE directs, Pete & Kelly execute LinkedIn outreach + verification under Bethany's supervision.",
            })
        if fp["total_placements_net"] > 0 or fp["recovered"] > 0:
            sections.append(_placements_table_section(fp, quarter_labels))
            target = target_by_provider.get(name)
            sections.append(_placements_chart_section(fp, quarter_labels, target))
            # Recovery Operation's CPP is "cost per recovered placement,"
            # not "cost per delivered placement" — the training-provider
            # threshold bands don't apply. Contract & Spend already
            # carries the relevant story.
            if fp["true_cpp"] > 0 and fp["category"] != "recovery":
                if _cpp_tone(fp["true_cpp"]) in ("watch", "critical"):
                    sections.append(_cost_analysis_verdict_section(fp))
                else:
                    sections.append(_cost_analysis_rows_section(fp))
        if related_items:
            sections.append(_action_items_section_for(related_items))

        entry = {
            "eyebrow": "Provider",
            "title": name,
            "summary": f"Full provider view · {fp.get('total_placements_net', 0)} reported + {fp.get('recovered', 0)} recovered",
            "sections": sections,
        }
        if fp["true_cpp"] > 0 and _cpp_tone(fp["true_cpp"]) == "critical":
            entry["status_chip"] = {
                "label": "Cost-per-placement in red band",
                "tone": "critical",
            }
        drills[f"provider:{name}"] = entry

    # ---------- Strategic partner drills (no placements but still budget lines) ----------
    for name, p in all_providers.items():
        if name in fp_by_provider or name in [r["provider"] for r in data["financial_performance"]]:
            continue
        if p.get("category") in ("strategic", "cfa_contractor"):
            drills[f"provider:{name}"] = {
                "eyebrow": "Strategic Partner" if p["category"] == "strategic" else "CFA Contractor",
                "title": name,
                "summary": p.get("notes", "")[:140],
                "sections": [{
                    "id": "contract_and_spend",
                    "type": "rows",
                    "title": "Contract & Spend",
                    "rows": [
                        {"label": "Amended budget", "value": f"${p['budget']:,.0f}"},
                        {"label": "Paid to date", "value": f"${p['qb_actual']:,.0f}"},
                        {"label": "Balance remaining", "value": f"${p['balance']:,.0f}"},
                        {"label": "Role", "value": "Engagement / support / convening — not placement-based"},
                    ],
                    "note": p.get("notes", ""),
                }],
            }

    # ---------- Budget category drills ----------
    for cat in data["summary"]["categories"]:
        key = f"category:{cat['name']}"
        rows = [
            {"label": "Budget", "value": f"${cat['budget']:,.0f}"},
            {"label": "Spent", "value": f"${cat['spent']:,.0f}"},
            {"label": "Remaining", "value": f"${cat['remaining']:,.0f}", "emphasize": True},
            {"label": "% used", "value": f"{cat['pct']:.1f}%"},
            {"label": "Monthly runway at current pace", "value": f"${cat['remaining'] / data['summary']['months_remaining']:,.0f}/mo"},
        ]
        if cat.get("prorated"):
            rows.append({"label": "⚠ Data quality", "value": "Pro-rated from $1,227,851 QB lump sum — real breakout pending from Krista"})

        # Every category drill gets the "Transaction detail pending" chip.
        # The old per-drill boilerplate note ("Once Class tracking and
        # production QB sync land...") was the same across seven drills —
        # promoted once, shown at the top of the panel.
        drills[key] = {
            "eyebrow": "Budget Category",
            "title": cat["name"],
            "summary": cat.get("note", f"${cat['remaining']:,.0f} remaining across {data['summary']['months_remaining']} months"),
            "status_chip": {
                "label": "Transaction detail pending",
                "tone": "watch",
            },
            "sections": [{
                "id": "current_status",
                "type": "rows",
                "title": "Current Status",
                "rows": rows,
            }],
        }

    # ---------- Decision drills ----------
    for i, item in enumerate(data["action_items"]):
        drills[f"decision:{i}"] = {
            "eyebrow": f"Action Item · {item['priority']}",
            "title": item["area"] if item["area"] else "(no area)",
            "summary": f"Owner: {item['owner']}",
            "sections": [{
                "id": "action_description",
                "type": "rows",
                "title": "Full action description",
                "rows": [{"label": "Action", "value": item["action"]}],
                "note": "Source: v3 Reconciliation Action Items sheet (3/27/26). Will update as Krista/Bethany work through items.",
            }],
        }

    # ---------- Audit dimension drills ----------
    # Dimension metadata + readiness come from extract_all's
    # audit_dimensions_from_engine fetch (step 2 cockpit-side).
    # "Open gaps" content comes from a per-dimension gap fetch
    # (step 4 cockpit-side, this block) via fetch_audit_gaps_from_engine.
    # See audit_readiness_tab_spec.md §v1.2.4 for the three-state
    # dimension-status contract and §v1.2.7 for the gap shape.
    #
    # TODO(v1.2 cockpit-side step 4 follow-up): add pytest on this
    # branch to cover the summary/readiness/gap rendering branches.
    # Tests deferred per scope; see integration_notes.md on
    # feature/compliance-engine-extract.
    engine_response = data.get("audit_dimensions_from_engine") or {}
    for dim in engine_response.get("dimensions", []):
        owner_display = display_name_for_role(dim.get("owner_role"))
        pct = dim.get("readiness_pct")
        status = dim.get("status", "placeholder")

        if status == "computed" and pct is not None:
            summary = f"{pct}% audit-ready · owner: {owner_display}"
            readiness_value: str | int = f"{pct}%"
        elif status == "computed":
            summary = f"Awaiting first scan · owner: {owner_display}"
            readiness_value = "—"
        else:
            summary = (
                f"Readiness measurement not yet available · owner: {owner_display}"
            )
            readiness_value = "—"

        gap_response = fetch_audit_gaps_from_engine(dim["id"])
        gap_section = _build_audit_gap_section(gap_response)

        drills[f"audit:{dim['id']}"] = {
            "eyebrow": "Audit Dimension",
            "title": dim.get("title", dim["id"]),
            "summary": summary,
            "sections": [
                {
                    "id": "auditor_perspective",
                    "type": "rows",
                    "title": "What auditors look for",
                    "rows": [{
                        "label": dim.get("what_auditors_look_for", ""),
                        "value": readiness_value,
                    }],
                },
                gap_section,
            ],
        }

    # ---------- Hero cell drills ----------
    # Four entries keyed by bare name (backbone, placements, reimbursement,
    # flags) to match the hero cells' data-drill attributes. Same rigid
    # {title, rows: [{label, value}]} schema as everything else — content
    # ported verbatim from the old heroDrillContent inline object in
    # cockpit_template.html.
    summary = data["summary"]
    placements = data["placements"]
    action_items = data["action_items"]
    recovered = data["recovered"]
    categories = summary["categories"]
    trailing_q1_total = sum(
        row["invoice"] for row in placements["q1_provider_actuals_breakdown"]
        if row["invoice"] is not None
    )
    high_priority_count = sum(
        1 for item in action_items if item["priority"] == "HIGH"
    )

    drills["backbone"] = {
        "eyebrow": "Backbone Runway",
        "title": f"${summary['backbone_runway_combined']:,.0f} across {summary['months_remaining']} months",
        "summary": "All CFA-side operations — staff, overhead, and recovery contractors",
        "sections": [
            {
                "id": "staff_overhead",
                "type": "rows",
                "title": f"Staff & overhead — ${summary['backbone_remaining']:,.0f} remaining",
                "rows": [
                    {"label": "Personnel — Salaries · ~$36,852/mo · ~6.0 months runway",
                     "value": f"${categories[3]['remaining']:,.0f}"},
                    {"label": "Personnel — Benefits · ~$5,815/mo",
                     "value": f"${categories[4]['remaining']:,.0f}"},
                    {"label": "Other Direct Costs · travel, comms, supplies",
                     "value": f"${categories[5]['remaining']:,.0f}"},
                    {"label": "Indirect Costs · de minimis 10%",
                     "value": f"${categories[6]['remaining']:,.0f}"},
                ],
            },
            {
                "id": "recovery_contractors",
                "type": "rows",
                "title": f"Recovery contractors — ${summary['cfa_contractor_remaining']:,.0f} remaining",
                "rows": [
                    {"label": "AI Engage · directs the recovery work",
                     "value": "$105,000"},
                    {"label": "Pete & Kelly Vargo · LinkedIn outreach + verification",
                     "value": "$105,217"},
                ],
            },
        ],
        "note": (
            "Verdict: At current burn (~$78k/month combined), backbone "
            "runway lands within ~$3k of the September 30 grant end. "
            "The $700k+ projected unspent in GJC Contractors is the only "
            "available buffer — moving even $200k via budget amendment "
            "buys real recovery-work runway."
        ),
    }

    drills["placements"] = {
        "eyebrow": "Placements",
        "title": f"{placements['confirmed_total']} confirmed of {placements['grant_goal']:,}",
        "summary": f"PIP threshold ({placements['pip_threshold']}) cleared on April 6 · 255 to grant goal",
        "sections": [
            {
                "id": "placements_sources",
                "type": "rows",
                "title": f"Where the {placements['confirmed_total']} came from",
                "rows": [
                    {"label": "Coalition reported placements · Q4 net of retractions",
                     "value": str(placements["coalition_reported"])},
                    {"label": "CFA verified Good Jobs · AIE + P&K LinkedIn recovery",
                     "value": str(placements["cfa_verified"])},
                    {"label": "Provider Q1 actuals · final GJC contributions",
                     "value": str(placements["q1_provider_actuals"])},
                ],
            },
        ],
        "note": (
            f"Live data. Synced from WJI TWC Candidate Tracking. "
            f"{recovered.get('total_validated', 0)} validated rows across "
            f"9 providers."
        ),
    }

    drills["reimbursement"] = {
        "eyebrow": "Cash Position",
        "title": f"${trailing_q1_total:,.0f} awaiting ESD reimbursement",
        "summary": "Q1 provider invoices paid by CFA · invoiced to ESD April 30",
        "sections": [
            {
                "id": "outstanding_receivable",
                "type": "rows",
                "title": "Outstanding receivable from ESD",
                "rows": [
                    {"label": f"{row['provider']} · {row['actual']} placements × ${row['rate']:,}",
                     "value": f"${row['invoice']:,}"}
                    for row in placements["q1_provider_actuals_breakdown"]
                    if row.get("invoice")
                ],
            },
        ],
        "note": (
            "Cycle timing. Invoice goes to ESD April 30 with the monthly "
            "advance request. Historical turnaround ~21 days, funds "
            "expected ~May 21. CFA must float this amount from operating "
            "cash for 3 weeks."
        ),
    }

    drills["flags"] = {
        "eyebrow": "Critical Flags",
        "title": f"{high_priority_count} items need decisions this week",
        # "status = open" removed — action_items has no status field.
        # The workflow (closed items are deleted, not flagged) keeps this
        # descriptively correct without enforcing it in code. Adding a
        # real status column is a larger conversation with Krista/Bethany.
        "summary": "Filtered to severity = HIGH",
        "sections": [
            {
                "id": "high_priority_items",
                "type": "rows",
                "title": "High-priority items from v3 reconciliation",
                "rows": [
                    {"label": f"{item['area']} · {item['owner']}",
                     "value": item["action"][:120]}
                    for item in action_items
                    if item["priority"] == "HIGH"
                ],
            },
        ],
        "note": (
            "Click any provider name throughout the cockpit to see their "
            "full detail view, including related action items."
        ),
    }

    # Fail loudly if any drill entry drifts from the schema.
    validate_registry(drills)

    return drills


def build_chart_data(data: dict) -> dict:
    """Pre-compute SVG coordinates for cockpit charts.

    Returns structured dict keyed by chart name with all geometry resolved
    so templates can just iterate and emit SVG elements. Kept separate from
    data extraction so visualizations can evolve without touching sources.
    """
    charts = {}

    # ---------- CHART 1: Budget Allocation stacked bar ----------
    # Single horizontal bar divided into the 7 categories (Training, Strategic,
    # CFA Contractors, Salaries, Benefits, Other, Indirect).
    # Each segment also shows spent-so-far as a darker inner portion.
    s = data["summary"]
    b = data["budget"]
    other_direct = b["travel"] + b["communications"] + b["supplies"] + b["other"]
    prorate = s["backbone_qb_paid"] / s["backbone_budget"]

    segments = [
        {"name": "GJC Training", "budget": s["training_budget"], "spent": s["training_paid"], "color": "#2E4D3E"},
        {"name": "GJC Strategic", "budget": s["strategic_budget"], "spent": s["strategic_paid"], "color": "#5C7A6B"},
        {"name": "CFA Contractors", "budget": s["cfa_contractor_budget"], "spent": s["cfa_contractor_paid"], "color": "#8B5E3C"},
        {"name": "Salaries", "budget": b["salaries"], "spent": b["salaries"] * prorate, "color": "#B8821B"},
        {"name": "Benefits", "budget": b["benefits"], "spent": b["benefits"] * prorate, "color": "#D4A847"},
        {"name": "Other Direct", "budget": other_direct, "spent": other_direct * prorate, "color": "#8C8A82"},
        {"name": "Indirect", "budget": b["indirect"], "spent": b["indirect"] * prorate, "color": "#C4B89C"},
    ]
    total_budget = sum(seg["budget"] for seg in segments)
    # Width resolution: use 1000 units, each segment gets width proportional to budget
    x = 0
    for seg in segments:
        seg["x"] = x
        seg["width"] = seg["budget"] / total_budget * 1000
        # spent pixels within this segment
        seg["spent_width"] = seg["spent"] / seg["budget"] * seg["width"] if seg["budget"] else 0
        seg["pct_used"] = seg["spent"] / seg["budget"] * 100 if seg["budget"] else 0
        x += seg["width"]
    charts["budget_allocation"] = {
        "segments": segments,
        "total_budget": total_budget,
        "total_spent": sum(seg["spent"] for seg in segments),
    }

    # ---------- CHART 2: Cumulative Placements Over Time ----------
    # One cumulative line for provider-reported, one for recovered, stacked.
    # Plus reference lines at 730 (PIP) and 1000 (goal).
    #
    # NOTE: quarterly_placements includes an "AI Engage" row with 256 placements
    # attributed in Q1'26. That same 256 is also represented via the recovery
    # data (by_provider). To avoid double-counting in the chart, we sum only
    # training providers for the "reported" line, and use the separate recovery
    # total (sourced from Bethany Validation sheet) for the overlay band.
    qp = [r for r in data["placements"]["quarterly_placements"]
          if r["provider"] != "AI Engage"]
    quarters = ["Q1'24", "Q2'24", "Q3'24", "Q4'24", "Q1'25", "Q2'25",
                "Q3'25", "Q4'25", "Q1'26"]
    # Sum per quarter across training providers only (provider-reported)
    reported_per_q = [0] * 9
    for row in qp:
        for i, n in enumerate(row["q"]):
            reported_per_q[i] += n
        reported_per_q[8] += row["q1_26_actual"] + row["q1_26_retraction"]
    reported_cum = []
    total = 0
    for n in reported_per_q:
        total += n
        reported_cum.append(total)
    # Recovered: all attributed at Q1'26 (that's when Bethany validated them)
    recovered_total = data["recovered"].get("total_validated", 255)
    recovered_cum = [0] * 8 + [recovered_total]

    # SVG dimensions
    w, h, pad_l, pad_r, pad_t, pad_b = 820, 260, 60, 40, 30, 40
    chart_w = w - pad_l - pad_r
    chart_h = h - pad_t - pad_b
    y_max = 1050  # enough headroom above 1000
    def x_at(i):
        return pad_l + (i / (len(quarters) - 1)) * chart_w
    def y_at(v):
        return pad_t + chart_h - (v / y_max) * chart_h

    # Build path strings
    reported_path_points = [(x_at(i), y_at(reported_cum[i])) for i in range(len(quarters))]
    combined_cum = [reported_cum[i] + recovered_cum[i] for i in range(len(quarters))]
    combined_path_points = [(x_at(i), y_at(combined_cum[i])) for i in range(len(quarters))]

    charts["cumulative_placements"] = {
        "width": w, "height": h,
        "pad_l": pad_l, "pad_t": pad_t, "pad_r": pad_r, "pad_b": pad_b,
        "chart_w": chart_w, "chart_h": chart_h,
        "quarters": quarters,
        "quarter_x": [x_at(i) for i in range(len(quarters))],
        "y_max": y_max,
        "y_pip": y_at(730),
        "y_goal": y_at(1000),
        "reported_cum": reported_cum,
        "recovered_cum": recovered_cum,
        "combined_cum": combined_cum,
        "reported_path": " ".join(f"{x:.1f},{y:.1f}" for x, y in reported_path_points),
        "combined_path": " ".join(f"{x:.1f},{y:.1f}" for x, y in combined_path_points),
        "reported_area_path": (
            f"M{pad_l},{y_at(0)} "
            + " ".join(f"L{x:.1f},{y:.1f}" for x, y in reported_path_points)
            + f" L{x_at(len(quarters)-1)},{y_at(0)} Z"
        ),
        "recovery_band_path": (
            " ".join(f"L{x:.1f},{y:.1f}" for x, y in combined_path_points)
            + " " + " ".join(f"L{x:.1f},{y:.1f}" for x, y in reversed(reported_path_points))
        ),
        "current_total": combined_cum[-1],
    }

    # ---------- CHART 3: Provider Contribution Stacked Bars ----------
    fp = data["financial_performance"]
    # Training providers + CFA Direct, excluding Recovery Operation (meta-row)
    bars = [r for r in fp if r["category"] in ("training", "cfa_direct")]
    # sort by true placements descending
    bars = sorted(bars, key=lambda r: -r["true_placements"])

    bar_w, bar_h = 600, 22
    bar_gap = 10
    max_plac = max(max((r["true_placements"] for r in bars), default=0),
                   max((r.get("target_override") or 0 for r in bars), default=0),
                   100)
    # Need target for each provider — pull from quarterly_placements
    target_map = {r["provider"]: r["target"]
                  for r in data["placements"]["quarterly_placements"]}
    max_plac = max(max_plac, max(target_map.values(), default=100))

    for bar in bars:
        target = target_map.get(bar["provider"], 0)
        bar["target"] = target
        bar["reported_width"] = bar["total_placements_net"] / max_plac * bar_w
        bar["recovered_width"] = bar["recovered"] / max_plac * bar_w
        bar["target_x"] = target / max_plac * bar_w
        # Tone for the % badge next to each bar. Moved out of Jinja
        # ternaries so all color-band logic lives in Python.
        if target > 0:
            pct_of_target = round(bar["true_placements"] / target * 100)
            bar["pct_of_target"] = pct_of_target
            bar["pct_tone"] = (
                "good" if pct_of_target >= 100
                else ("watch" if pct_of_target >= 75 else "critical")
            )
        else:
            bar["pct_of_target"] = 0
            bar["pct_tone"] = "neutral"

    charts["provider_contribution"] = {
        "bars": bars,
        "bar_width": bar_w,
        "bar_height": bar_h,
        "bar_gap": bar_gap,
        "max_plac": max_plac,
        "label_width": 160,
        "total_height": len(bars) * (bar_h + bar_gap),
    }

    # ---------- CHART 4: True CPP Ranking ----------
    # Sort providers (only those with CPP > 0) by true_cpp ascending
    cpp_rows = [r for r in fp if r["true_cpp"] > 0]
    cpp_rows = sorted(cpp_rows, key=lambda r: r["true_cpp"])
    max_cpp = max((r["true_cpp"] for r in cpp_rows), default=10000)
    # Round max up to next $2k for clean axis
    max_cpp_axis = ((int(max_cpp) // 2000) + 1) * 2000

    cpp_bar_w = 520
    for r in cpp_rows:
        r["cpp_width"] = r["true_cpp"] / max_cpp_axis * cpp_bar_w
        r["reported_cpp_width"] = r["cpp"] / max_cpp_axis * cpp_bar_w
        if r["true_cpp"] <= 2500:
            r["cpp_color"] = "#1F6B5C"  # good
        elif r["true_cpp"] <= 4000:
            r["cpp_color"] = "#B8821B"  # watch
        else:
            r["cpp_color"] = "#B8462E"  # critical

    charts["cpp_ranking"] = {
        "rows": cpp_rows,
        "bar_width": cpp_bar_w,
        "max_cpp_axis": max_cpp_axis,
        "label_width": 200,
        "bar_height": 22,
        "bar_gap": 8,
        "threshold_good": 2500 / max_cpp_axis * cpp_bar_w,
        "threshold_watch": 4000 / max_cpp_axis * cpp_bar_w,
    }

    return charts


def extract_recovered_placements(twc_path: Path) -> dict:
    """Read WJI TWC Candidate Tracking — Bethany Validation sheet.

    Returns per-provider count of Validated placements (the recovered Good Jobs)
    plus status totals. In production this comes from the SharePoint sync.
    """
    from collections import Counter, defaultdict
    if twc_path is None or not twc_path.exists():
        return {"by_provider": {}, "by_status": {}, "total_validated": 0,
                "available": False}

    wb = load_workbook(twc_path, data_only=True)
    ws = wb["Bethany Validation"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Map raw provider names in TWC to cockpit-canonical names
    name_map = {
        "Ada": "Ada",
        "CodeDay/MinT": "Code Day",
        "Computing For All": "CFA Direct",
        "NCESD": "NCESD",
        "PNW CCG": "PNW CCG",
        "Per Scholas": "Per Scholas",
        "Riipen": "Riipen",
        "St Martins - Washington Vets 2 Tech": "Vets2Tech",
        "Year Up": "Year Up",
    }

    by_provider = defaultdict(int)
    by_status = Counter()
    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue
        provider_raw = row[4]
        status = row[15]
        by_status[status if status else "(blank)"] += 1
        if status == "Validated":
            canon = name_map.get(provider_raw, provider_raw or "(unattributed)")
            by_provider[canon] += 1

    return {
        "by_provider": dict(by_provider),
        "by_status": dict(by_status),
        "total_validated": by_status.get("Validated", 0),
        "available": True,
    }


def _to_float(v) -> float:
    """Best-effort parse to float. Strings with commas, '—', None all OK."""
    if v is None or v == "—":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("$", "").strip()
    # Handle multi-line cells like "149,323 gross\n67,020 net" — take first number
    if "\n" in s:
        s = s.split("\n")[0].strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def build_financial_performance(cpp: dict, placements: dict, providers: dict,
                                recovered: dict | None = None) -> list:
    """Combine quarterly payments + quarterly placements + recovery counts per provider.

    Returns one row per placement-producing entity with:
    - quarterly payments Q1'24 → Q4'25 (from Cost Per Placement spreadsheet)
    - Q1'26 final invoice amount (from placements breakdown, may be pending)
    - total paid (cash basis from CPP)
    - total net placements as reported by provider (from quarterly history)
    - recovered placements (validated by CFA via LinkedIn outreach)
    - true placements = reported + recovered
    - reported CPP = paid / reported
    - true CPP = paid / (reported + recovered)
    """
    recovered = recovered or {"by_provider": {}, "available": False}
    recovered_by_provider = recovered.get("by_provider", {})
    # Map provider names between the two data sources (cost_per_placement uses
    # slightly different names than quarterly_placements)
    name_map = {
        "Ada": "Ada",
        "Code Day X Mint": "Code Day",
        "Per Scholas": "Per Scholas",
        "PNW CCG": "PNW CCG",
        "Riipen": "Riipen",
        "Year Up": "Year Up",
        "Vets2Tech": "Vets2Tech",
        "NCESD": "NCESD",
    }
    placements_by_name = {row["provider"]: row for row in placements["quarterly_placements"]}
    q1_26_invoice_by_provider = {
        row["provider"]: row.get("invoice")
        for row in placements["q1_provider_actuals_breakdown"]
    }
    # Some Q1 invoice keys differ from quarterly_placements keys
    q1_invoice_aliases = {
        "Ada Developers": "Ada",
        "Code Day / MinT": "Code Day",
    }

    rows = []
    for cpp_provider in cpp["providers"]:
        cpp_name = cpp_provider["name"]
        match_name = name_map.get(cpp_name, cpp_name)
        placement_row = placements_by_name.get(match_name)
        if not placement_row:
            continue

        # Find Q1'26 invoice (may be under a different name in q1 breakdown)
        q1_invoice = None
        for q1_name, alias in {**{k: k for k in q1_26_invoice_by_provider}, **q1_invoice_aliases}.items():
            if alias == match_name:
                q1_invoice = q1_26_invoice_by_provider.get(q1_name)
                break

        net_placements = (sum(placement_row["q"])
                          + placement_row["q1_26_actual"]
                          + placement_row["q1_26_retraction"])
        cpp_value = (cpp_provider["total_paid"] / net_placements
                     if net_placements > 0 else 0)
        recovered_count = recovered_by_provider.get(match_name, 0)
        true_placements = net_placements + recovered_count
        true_cpp = (cpp_provider["total_paid"] / true_placements
                    if true_placements > 0 else 0)

        rows.append({
            "provider": match_name,
            "category": "training",
            "quarterly_payments": cpp_provider["quarterly"],
            "quarterly_placements": placement_row["q"],
            "q1_26_invoice": q1_invoice,
            "q1_26_placements": placement_row["q1_26_actual"],
            "q1_26_retraction": placement_row["q1_26_retraction"],
            "total_paid": cpp_provider["total_paid"],
            "total_placements_net": net_placements,
            "recovered": recovered_count,
            "true_placements": true_placements,
            "cpp": cpp_value,
            "true_cpp": true_cpp,
            # Tones pre-computed in Python so the template never runs a
            # color-band ternary.
            "cpp_tone": _cpp_tone(cpp_value),
            "true_cpp_tone": _cpp_tone(true_cpp),
        })

    # Add Apprenti (training but not in CPP file due to no payments yet)
    apprenti = placements_by_name.get("Apprenti")
    if apprenti:
        recovered_count = recovered_by_provider.get("Apprenti", 0)
        rows.append({
            "provider": "Apprenti",
            "category": "training",
            "quarterly_payments": [0]*8,
            "quarterly_placements": apprenti["q"],
            "q1_26_invoice": None,
            "q1_26_placements": apprenti["q1_26_actual"],
            "q1_26_retraction": apprenti["q1_26_retraction"],
            "total_paid": 0,
            "total_placements_net": 0,
            "recovered": recovered_count,
            "true_placements": recovered_count,
            "cpp": 0,
            "true_cpp": 0,
            "cpp_tone": _cpp_tone(0),
            "true_cpp_tone": _cpp_tone(0),
        })

    # CFA Direct row — placements found by recovery work that aren't traceable
    # to a coalition training provider (probably P&K's organic outreach)
    cfa_direct = recovered_by_provider.get("CFA Direct", 0)
    if cfa_direct > 0:
        rows.append({
            "provider": "CFA Direct (no training provider)",
            "category": "cfa_direct",
            "quarterly_payments": [0]*8,
            "quarterly_placements": [0]*8,
            "q1_26_invoice": None,
            "q1_26_placements": 0,
            "q1_26_retraction": 0,
            "total_paid": 0,
            "total_placements_net": 0,
            "recovered": cfa_direct,
            "true_placements": cfa_direct,
            "cpp": 0,
            "true_cpp": 0,
            "cpp_tone": _cpp_tone(0),
            "true_cpp_tone": _cpp_tone(0),
        })

    # Recovery Operation row (AI Engage + Pete & Kelly combined)
    aie_paid = next((p["qb_actual"] for p in providers["cfa_contractors"]
                     if p["name"].startswith("AI Engage")), 0)
    pk_paid = next((p["qb_actual"] for p in providers["cfa_contractors"]
                    if "Pete" in p["name"] or "Vargo" in p["name"]), 0)
    recovery_paid = aie_paid + pk_paid
    recovery_placements = recovered.get("total_validated", 256)
    recovery_cpp = recovery_paid / recovery_placements if recovery_placements else 0
    rows.append({
        "provider": "Recovery Operation (AIE + P&K)",
        "category": "recovery",
        "quarterly_payments": [0]*8,
        "quarterly_placements": [0]*8,
        "q1_26_invoice": None,
        "q1_26_placements": recovery_placements,
        "q1_26_retraction": 0,
        "total_paid": recovery_paid,
        "total_placements_net": recovery_placements,
        "recovered": 0,  # recovery operation IS the source of recovered, doesn't get re-credited
        "true_placements": recovery_placements,
        "cpp": recovery_cpp,
        "true_cpp": recovery_cpp,
        "cpp_tone": _cpp_tone(recovery_cpp),
        "true_cpp_tone": _cpp_tone(recovery_cpp),
    })

    return rows



    """Best-effort parse to float. Strings with commas, '—', None all OK."""
    if v is None or v == "—":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("$", "").strip()
    # Handle multi-line cells like "149,323 gross\n67,020 net" — take first number
    if "\n" in s:
        s = s.split("\n")[0].strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


DEFAULT_DATA_DIR = Path(__file__).resolve().parent / "fixtures"


def resolve_data_dir(project_dir: "Path | str | None" = None) -> Path:
    """Resolve the spreadsheet source directory.

    Precedence: explicit argument > COCKPIT_DATA_DIR env var >
    `agents/finance/design/fixtures/` next to this module.
    """
    import os
    if project_dir is not None:
        return Path(project_dir)
    env = os.environ.get("COCKPIT_DATA_DIR")
    if env:
        return Path(env)
    return DEFAULT_DATA_DIR


def _find_fixture(project_dir: Path, patterns: list,
                  *, required: bool = True):
    """Return the first xlsx in `project_dir` matching any glob pattern.

    Tolerates filename drift — hyphens vs underscores, spaces, ampersands,
    trailing "(1)" — so callers don't have to hardcode exact names.
    """
    for pat in patterns:
        matches = sorted(project_dir.glob(pat))
        if matches:
            return matches[0]
    if required:
        raise FileNotFoundError(
            f"No fixture in {project_dir} matched any of: {patterns}"
        )
    return None


def extract_all(project_dir=None) -> dict:
    """Top-level entry: read all source spreadsheets, return structured data.

    `project_dir` defaults to agents/finance/design/fixtures/. Override via
    the `COCKPIT_DATA_DIR` environment variable or by passing a path.
    """
    project_dir = resolve_data_dir(project_dir)
    recon = _find_fixture(project_dir, [
        "*Provider_Reconciliation*.xlsx",
        "*Provider Reconciliation*.xlsx",
    ])
    cpp = _find_fixture(project_dir, [
        "*Cost_Per_Placement*.xlsx",
        "*Cost Per Placement*.xlsx",
    ])
    exhb = _find_fixture(project_dir, [
        "*Exh_B*.xlsx",
        "*Exh B*.xlsx",
    ])
    twc = _find_fixture(project_dir, [
        "*TWC_Candidate_Tracking*.xlsx",
        "*TWC Candidate Tracking*.xlsx",
        "WJI_TWC*.xlsx",
        "WJI TWC*.xlsx",
    ], required=False)

    providers = extract_providers(recon)
    action_items = extract_action_items(recon)
    cost_per_placement = extract_cost_per_placement(cpp)
    budget = extract_budget_categories(exhb)
    recovered = extract_recovered_placements(twc)
    summary = compute_summary(providers, cost_per_placement, budget)

    # Placement data from CLAUDE.md / dashboard — not in spreadsheets
    # In production this comes from WJI TWC SharePoint sync
    placements = {
        "confirmed_total": 745,
        "pip_threshold": 730,
        "grant_goal": 1000,
        "coalition_reported": 421,
        "cfa_verified": 256,
        "q1_provider_actuals": 68,
        "vets2tech_q2_guaranteed": 24,
        "apprenti_expected_low": 18,
        "apprenti_expected_high": 32,
        "recovery_target": 231,
        "total_participants": 1785,
        "confirmed_plus_guaranteed": 769,
        "linkedin_unreachable": 285,
        "reachable_pool": 731,
        "live_synced_minutes_ago": 3,
        "q1_provider_actuals_breakdown": [
            {"provider": "Ada Developers", "expected": "13+", "actual": 28,
             "rate": 2500, "invoice": 70000, "variance_color": "good"},
            {"provider": "Vets2Tech", "expected": "24", "actual": 12,
             "rate": 2500, "invoice": 30000, "variance_color": "critical"},
            {"provider": "Year Up", "expected": "5", "actual": 11,
             "rate": 2623, "invoice": 28853, "variance_color": "good"},
            {"provider": "Per Scholas", "expected": "11", "actual": 9,
             "rate": 3443, "invoice": 30987, "variance_color": "neutral"},
            {"provider": "Code Day / MinT", "expected": "14", "actual": 8,
             "rate": 3222, "invoice": 25776, "variance_color": "watch"},
            {"provider": "Apprenti", "expected": "32", "actual": None,
             "rate": 2500, "invoice": None, "variance_color": "neutral"},
        ],
        # Quarterly placement matrix — Q1'24 through Q4'25, plus Q1'26 actuals
        # and retractions. Source: CLAUDE.md (derived from WJI TWC tracking).
        # In production this comes from the WSAC Partner Data Outcomes sync.
        "quarterly_placements": [
            {"provider": "Year Up",     "q": [16, 15,  6,  5, 18, 14,  7,  5], "total_through_q4_25": 86, "q1_26_actual": 11, "q1_26_retraction": 0,  "target": 81},
            {"provider": "Ada",         "q": [29,  1, 28,  1, 19,  0,  2,  2], "total_through_q4_25": 82, "q1_26_actual": 28, "q1_26_retraction": 0,  "target": 136},
            {"provider": "Vets2Tech",   "q": [ 0, 11,  8, 11, 14, 15,  4,  5], "total_through_q4_25": 68, "q1_26_actual": 12, "q1_26_retraction": 0,  "target": 105},
            {"provider": "Code Day",    "q": [ 1, 14,  6,  8,  3, 24,  8,  4], "total_through_q4_25": 68, "q1_26_actual":  8, "q1_26_retraction": 0,  "target": 78},
            {"provider": "NCESD",       "q": [ 0,  2,  5,  5,  8, 10, 45,  0], "total_through_q4_25": 75, "q1_26_actual":  0, "q1_26_retraction": -19, "target": 81},
            {"provider": "Per Scholas", "q": [ 3,  3,  4,  0, 11,  3,  7,  2], "total_through_q4_25": 33, "q1_26_actual":  9, "q1_26_retraction": 0,  "target": 44},
            {"provider": "PNW CCG",     "q": [ 0,  0,  2,  0, 13,  5,  1,  5], "total_through_q4_25": 26, "q1_26_actual":  0, "q1_26_retraction": 0,  "target": 27},
            {"provider": "Riipen",      "q": [ 0,  0,  1,  1,  1,  1,  0,  0], "total_through_q4_25":  4, "q1_26_actual":  0, "q1_26_retraction": 0,  "target": 25},
            {"provider": "Apprenti",    "q": [ 0,  0,  0,  0,  0,  0,  0,  0], "total_through_q4_25":  0, "q1_26_actual":  0, "q1_26_retraction": 0,  "target": 17},
            # AI Engage attributed with 256 placements recovered in Q1 2026 via LinkedIn outreach.
            # No contractual per-placement target — recovery work is deliverable-shaped.
            {"provider": "AI Engage",   "q": [ 0,  0,  0,  0,  0,  0,  0,  0], "total_through_q4_25":  0, "q1_26_actual": 256, "q1_26_retraction": 0,  "target": 256},
        ],
        "quarter_labels": ["Q1'24", "Q2'24", "Q3'24", "Q4'24", "Q1'25", "Q2'25", "Q3'25", "Q4'25"],
    }

    # Pre-compute net / pct / pct_tone on every quarterly placement row so
    # the template never runs a band ternary for progress color.
    for row in placements["quarterly_placements"]:
        net = row["total_through_q4_25"] + row["q1_26_actual"] + row["q1_26_retraction"]
        pct = (net / row["target"] * 100) if row["target"] > 0 else 0
        row["net"] = net
        row["pct"] = pct
        row["pct_tone"] = (
            "good" if pct >= 100
            else ("watch" if pct >= 75 else "critical")
        )

    result = {
        "providers": providers,
        "action_items": action_items,
        "cost_per_placement": cost_per_placement,
        "budget": budget,
        "summary": summary,
        "placements": placements,
        "recovered": recovered,
        "financial_performance": build_financial_performance(
            cost_per_placement, placements, providers, recovered
        ),
    }
    # Fetch audit dimensions from the compliance engine once per
    # data-source refresh. Both build_drills (below) and
    # cockpit_api._tab_audit read from this key. If the engine is
    # unreachable the fetch returns a degraded-state marker; the
    # downstream consumers render dimension rows in a clearly
    # "unavailable" state rather than hiding the issue.
    result["audit_dimensions_from_engine"] = fetch_audit_dimensions_from_engine()
    result["audit_activity_from_engine"] = fetch_audit_activity_from_engine()
    result["charts"] = build_chart_data(result)
    result["drills"] = build_drills(result)
    return result


if __name__ == "__main__":
    import json
    data = extract_all()
    # Pretty-print summary for sanity check
    print(json.dumps(data["summary"], indent=2, default=str))
    print(f"\nProviders found: " + ", ".join(
        f"{k}={len(v)}" for k, v in data["providers"].items()))
    print(f"Action items: {len(data['action_items'])}")
    print(f"CPP providers: {len(data['cost_per_placement']['providers'])}")
