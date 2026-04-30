"""Personnel & Contractors view — data model + parser + reconciliation.

Implements the data layer for the Personnel & Contractors sub-section under
the Finance Cockpit's Budget & Burn tab. See
`agents/finance/design/personnel_contractors_view_spec.md` for the spec
this module satisfies.

INPUT MECHANISM (decision recorded in the spec)
-----------------------------------------------
Excel workbook at `agents/finance/design/fixtures/K8341_Personnel_and_Contractors.xlsx`.
Matches the existing five-Excel pipeline (cockpit_data.py reads Exhibit B,
Provider Reconciliation, Cost Per Placement, etc. from the same fixtures
directory). Krista is a bookkeeper — Excel is her native medium, and the
quarterly update workflow (replace projected → actual at quarter close) is
naturally a row edit in a spreadsheet. Each commit of the .xlsx becomes the
audit trail.

Three sheets:
  People       — one row per grant-funded person (id, identity, budget
                 allocation, rate)
  Actuals      — one row per person-quarter payment that has happened
  Projections  — one row per person-quarter projection through grant end
  _README      — free-form schema documentation for Krista

The parser also writes a JSON snapshot
(`K8341_Personnel_and_Contractors.snapshot.json`) sibling to the workbook
at every extract, so PR diffs can show meaningful textual deltas of the
latest data rather than opaque .xlsx blobs.

QUARTERLY UPDATE WORKFLOW (for Krista)
--------------------------------------
At each quarter close:
  1. Open `K8341_Personnel_and_Contractors.xlsx`.
  2. On the Actuals sheet, add one row per person for the closed quarter
     with amount_paid (sourced from QB or payroll) and source.
  3. On the Projections sheet, REMOVE the projection row for the
     just-closed quarter (it's now an actual), and update remaining-quarter
     projections if your understanding of the trajectory has changed.
  4. Save and commit. The cockpit re-reads on next refresh.

For mid-quarter changes (new contractor, rate change, contract amendment),
edit the People sheet directly and commit.

HONESTY DISCIPLINE (from the spec)
----------------------------------
- A row missing required fields is flagged `documentation_incomplete=True`,
  not silently filled.
- Budget-line rollups that don't reconcile to Amendment 1 produce an entry
  in `reconciliation_warnings` — never silently absorbed.
- A person with zero projections through grant end is flagged as
  `projections_missing=True` rather than treated as projecting $0.
"""
from __future__ import annotations

from dataclasses import dataclass, field, asdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Optional
import json

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Reference values from Amendment 1 (approved Nov 2025)
# ---------------------------------------------------------------------------

# Per-line totals as documented in personnel_contractors_view_spec.md §
# "Reconciliation reference (Amendment 1)". GJC strategic-partner totals
# can be sourced at extract time from cockpit_data.extract_budget_categories
# if/when needed; for now the three lines below are the hard reconciliation
# constraint.
AMENDMENT_1_TOTALS: dict[str, float] = {
    "personnel_salaries": 1_097_662.41,
    "personnel_benefits": 173_169.94,
    "cfa_contractors": 1_020_823.40,
    # "gjc_contractors_strategic": TBD per Amendment 1 — extend when data lands
}

# Display labels — one place to change copy.
BUDGET_LINE_LABELS: dict[str, str] = {
    "personnel_salaries": "Personnel — Salaries",
    "personnel_benefits": "Personnel — Benefits",
    "cfa_contractors": "CFA Contractors",
    "gjc_contractors_strategic": "GJC Contractors — Strategic Partners",
}

# Reconciliation tolerance: differences smaller than this are considered
# rounding noise, not drift. Matches typical cents-precision payroll math.
RECONCILIATION_TOLERANCE_USD = 0.01

# Required fields per row (see "honesty discipline" above).
REQUIRED_PERSON_FIELDS: tuple[str, ...] = (
    "id", "name", "role", "engagement_type",
    "budget_line", "amended_budget_total",
    "rate_amount", "rate_unit", "rate_basis", "rate_effective_date",
    "start_date",
)

VALID_ENGAGEMENT_TYPES: frozenset[str] = frozenset({
    "employee", "contractor", "subcontractor",
})

VALID_RATE_UNITS: frozenset[str] = frozenset({
    "hourly", "daily", "weekly", "monthly", "annual",
    "fixed_fee", "milestone",
})

VALID_BUDGET_LINES: frozenset[str] = frozenset(BUDGET_LINE_LABELS.keys())

VALID_ACTUAL_SOURCES: frozenset[str] = frozenset({
    "qb", "payroll", "manual_entry",
})


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class QuarterlyActual:
    """One per (person, quarter) — a payment that has already happened."""
    quarter: str  # "YYYY-QN", e.g., "2024-Q1"
    amount_paid: float
    source: str  # qb | payroll | manual_entry
    qb_vendor_name: Optional[str] = None


@dataclass
class QuarterlyProjection:
    """One per (person, remaining quarter) — a projected payment."""
    quarter: str
    projected_amount: float
    projection_basis: str  # short note explaining the projection method


@dataclass
class Person:
    """One grant-funded person. Identity + budget allocation + rate +
    quarterly actuals + quarterly projections + computed fields."""
    # Identity
    id: str
    name: str
    role: str
    engagement_type: str
    vendor_legal_entity: Optional[str]
    start_date: Optional[date]
    end_date: Optional[date]  # None = through grant end

    # Budget allocation
    budget_line: str
    amended_budget_total: float

    # Rate
    rate_amount: float
    rate_unit: str
    rate_basis: str
    rate_effective_date: Optional[date]

    # Quarterly tables
    actuals: list[QuarterlyActual] = field(default_factory=list)
    projections: list[QuarterlyProjection] = field(default_factory=list)

    # Honesty flags (set by the parser/computer, not by the input file)
    missing_required_fields: list[str] = field(default_factory=list)
    projections_missing: bool = False

    # Computed fields (filled in by compute_derived())
    paid_to_date: float = 0.0
    projected_total_remaining: float = 0.0
    total_committed: float = 0.0
    variance_vs_amended: float = 0.0
    variance_pct: float = 0.0
    amended_budget_remaining_periods: float = 0.0

    @property
    def documentation_incomplete(self) -> bool:
        return bool(self.missing_required_fields)

    @property
    def drill_key(self) -> str:
        """Cockpit drill key, e.g., 'person:ritu-bahl'."""
        return f"person:{self.id}"


@dataclass
class BudgetLineRollup:
    """Per-budget-line totals, computed from the people in that line."""
    budget_line: str  # one of VALID_BUDGET_LINES
    label: str  # display label
    person_count: int
    amended_budget_total: float
    paid_to_date: float
    projected_total_remaining: float
    total_committed: float
    variance_vs_amended: float
    amendment_1_reference: Optional[float]
    reconciliation_delta: Optional[float]  # amended_budget_total - amendment_1_reference
    reconciles: Optional[bool]  # None = no reference available


@dataclass
class ReconciliationWarning:
    """One issue worth surfacing prominently in the rendered output."""
    level: str  # "error" | "warning" | "info"
    budget_line: Optional[str]  # None = global
    message: str


@dataclass
class PersonnelExtract:
    """Top-level result returned to the cockpit pipeline."""
    people: list[Person]
    rollups: list[BudgetLineRollup]
    distinct_person_count: int  # de-duped by name (an employee on Salaries
                                # + Benefits is one person, not two)
    summary_paid_to_date: float
    summary_total_committed: float
    summary_variance_vs_amended: float
    reconciliation_warnings: list[ReconciliationWarning]
    extracted_at: str  # ISO-8601 timestamp
    source_workbook: Optional[str]  # absolute path to the .xlsx, for trace


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------

PEOPLE_SHEET = "People"
ACTUALS_SHEET = "Actuals"
PROJECTIONS_SHEET = "Projections"


def _coerce_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _coerce_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        v = value.strip()
        return v or None
    return str(value).strip() or None


def _coerce_date(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def _row_dict(headers: list[str], row: tuple) -> dict:
    """Convert a row tuple to a dict keyed by lower-snake-cased header."""
    return {h: v for h, v in zip(headers, row)}


def parse_workbook(workbook_path: str | Path) -> tuple[list[Person], list[ReconciliationWarning]]:
    """Read the personnel workbook and return (people, parse_warnings).

    Each Person comes back fully populated EXCEPT computed fields — those
    are filled in by `compute_derived()`. Per-row `missing_required_fields`
    is populated here.

    Parse warnings cover structural problems (missing sheets, unknown
    enum values, orphan actual/projection rows). They are NOT row-level
    "documentation incomplete" — those are tracked on the Person itself.
    """
    path = Path(workbook_path)
    warnings: list[ReconciliationWarning] = []

    if not path.exists():
        warnings.append(ReconciliationWarning(
            level="error",
            budget_line=None,
            message=(
                f"Personnel workbook not found at {path}. "
                "Awaiting initial data population — see the Personnel & "
                "Contractors spec."
            ),
        ))
        return [], warnings

    wb = load_workbook(path, data_only=True, read_only=True)
    sheet_names = set(wb.sheetnames)

    for required in (PEOPLE_SHEET, ACTUALS_SHEET, PROJECTIONS_SHEET):
        if required not in sheet_names:
            warnings.append(ReconciliationWarning(
                level="error",
                budget_line=None,
                message=f"Required sheet '{required}' missing from workbook.",
            ))
    if any(w.level == "error" for w in warnings):
        return [], warnings

    # ---- People sheet ----
    people_by_id: dict[str, Person] = {}
    people_sheet = wb[PEOPLE_SHEET]
    rows = list(people_sheet.iter_rows(values_only=True))
    if not rows:
        return [], warnings
    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]

    for raw in rows[1:]:
        if all(c is None or c == "" for c in raw):
            continue  # skip blank rows
        d = _row_dict(headers, raw)
        pid = _coerce_str(d.get("id"))
        if not pid:
            warnings.append(ReconciliationWarning(
                level="warning",
                budget_line=None,
                message=f"People row missing id, skipping: {d}",
            ))
            continue
        if pid in people_by_id:
            warnings.append(ReconciliationWarning(
                level="warning",
                budget_line=None,
                message=f"Duplicate person id {pid!r}; later row wins.",
            ))

        engagement = _coerce_str(d.get("engagement_type"))
        if engagement and engagement not in VALID_ENGAGEMENT_TYPES:
            warnings.append(ReconciliationWarning(
                level="warning",
                budget_line=None,
                message=(
                    f"Person {pid!r} has unknown engagement_type "
                    f"{engagement!r}; expected one of "
                    f"{sorted(VALID_ENGAGEMENT_TYPES)}."
                ),
            ))
        rate_unit = _coerce_str(d.get("rate_unit"))
        if rate_unit and rate_unit not in VALID_RATE_UNITS:
            warnings.append(ReconciliationWarning(
                level="warning",
                budget_line=None,
                message=(
                    f"Person {pid!r} has unknown rate_unit "
                    f"{rate_unit!r}; expected one of "
                    f"{sorted(VALID_RATE_UNITS)}."
                ),
            ))
        budget_line = _coerce_str(d.get("budget_line"))
        if budget_line and budget_line not in VALID_BUDGET_LINES:
            warnings.append(ReconciliationWarning(
                level="warning",
                budget_line=budget_line,
                message=(
                    f"Person {pid!r} has unknown budget_line "
                    f"{budget_line!r}; expected one of "
                    f"{sorted(VALID_BUDGET_LINES)}."
                ),
            ))

        person = Person(
            id=pid,
            name=_coerce_str(d.get("name")) or "",
            role=_coerce_str(d.get("role")) or "",
            engagement_type=engagement or "",
            vendor_legal_entity=_coerce_str(d.get("vendor_legal_entity")),
            start_date=_coerce_date(d.get("start_date")),
            end_date=_coerce_date(d.get("end_date")),
            budget_line=budget_line or "",
            amended_budget_total=_coerce_float(d.get("amended_budget_total")) or 0.0,
            rate_amount=_coerce_float(d.get("rate_amount")) or 0.0,
            rate_unit=rate_unit or "",
            rate_basis=_coerce_str(d.get("rate_basis")) or "",
            rate_effective_date=_coerce_date(d.get("rate_effective_date")),
        )
        # Honesty: which required fields are missing/blank?
        missing: list[str] = []
        for f_name in REQUIRED_PERSON_FIELDS:
            val = getattr(person, f_name, None)
            if val is None or (isinstance(val, str) and not val) or (isinstance(val, float) and val == 0.0 and f_name != "amended_budget_total"):
                # amended_budget_total of 0 is plausible (a person with
                # zero allocation flagged for review); other zeros are not.
                missing.append(f_name)
            elif isinstance(val, float) and f_name == "amended_budget_total" and val < 0:
                missing.append(f_name)
        person.missing_required_fields = missing
        people_by_id[pid] = person

    # ---- Actuals sheet ----
    actuals_sheet = wb[ACTUALS_SHEET]
    rows = list(actuals_sheet.iter_rows(values_only=True))
    if rows:
        a_headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
        for raw in rows[1:]:
            if all(c is None or c == "" for c in raw):
                continue
            d = _row_dict(a_headers, raw)
            pid = _coerce_str(d.get("person_id"))
            if not pid:
                warnings.append(ReconciliationWarning(
                    level="warning", budget_line=None,
                    message=f"Actuals row missing person_id, skipping: {d}",
                ))
                continue
            if pid not in people_by_id:
                warnings.append(ReconciliationWarning(
                    level="warning", budget_line=None,
                    message=f"Actuals row references unknown person_id {pid!r}; skipping.",
                ))
                continue
            quarter = _coerce_str(d.get("quarter"))
            amount = _coerce_float(d.get("amount_paid"))
            source = _coerce_str(d.get("source"))
            if not quarter or amount is None or not source:
                warnings.append(ReconciliationWarning(
                    level="warning", budget_line=None,
                    message=(
                        f"Actuals row for {pid!r} missing quarter / "
                        f"amount_paid / source; skipping."
                    ),
                ))
                continue
            if source not in VALID_ACTUAL_SOURCES:
                warnings.append(ReconciliationWarning(
                    level="warning", budget_line=None,
                    message=(
                        f"Actuals row for {pid!r} has unknown source "
                        f"{source!r}; accepting."
                    ),
                ))
            people_by_id[pid].actuals.append(QuarterlyActual(
                quarter=quarter,
                amount_paid=amount,
                source=source,
                qb_vendor_name=_coerce_str(d.get("qb_vendor_name")),
            ))

    # ---- Projections sheet ----
    proj_sheet = wb[PROJECTIONS_SHEET]
    rows = list(proj_sheet.iter_rows(values_only=True))
    if rows:
        p_headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
        for raw in rows[1:]:
            if all(c is None or c == "" for c in raw):
                continue
            d = _row_dict(p_headers, raw)
            pid = _coerce_str(d.get("person_id"))
            if not pid:
                warnings.append(ReconciliationWarning(
                    level="warning", budget_line=None,
                    message=f"Projections row missing person_id, skipping: {d}",
                ))
                continue
            if pid not in people_by_id:
                warnings.append(ReconciliationWarning(
                    level="warning", budget_line=None,
                    message=f"Projections row references unknown person_id {pid!r}; skipping.",
                ))
                continue
            quarter = _coerce_str(d.get("quarter"))
            amount = _coerce_float(d.get("projected_amount"))
            basis = _coerce_str(d.get("projection_basis"))
            if not quarter or amount is None:
                warnings.append(ReconciliationWarning(
                    level="warning", budget_line=None,
                    message=(
                        f"Projections row for {pid!r} missing quarter "
                        f"or projected_amount; skipping."
                    ),
                ))
                continue
            people_by_id[pid].projections.append(QuarterlyProjection(
                quarter=quarter,
                projected_amount=amount,
                projection_basis=basis or "",
            ))

    return list(people_by_id.values()), warnings


# ---------------------------------------------------------------------------
# Computed fields + rollups + reconciliation
# ---------------------------------------------------------------------------

def compute_derived(person: Person) -> None:
    """Fill in the computed fields on a Person in place."""
    person.paid_to_date = sum(a.amount_paid for a in person.actuals)
    person.projected_total_remaining = sum(p.projected_amount for p in person.projections)
    person.total_committed = person.paid_to_date + person.projected_total_remaining
    person.variance_vs_amended = person.amended_budget_total - person.total_committed
    person.variance_pct = (
        person.variance_vs_amended / person.amended_budget_total
        if person.amended_budget_total else 0.0
    )
    person.amended_budget_remaining_periods = (
        person.amended_budget_total - person.paid_to_date
    )
    person.projections_missing = (not person.projections) and (
        person.amended_budget_remaining_periods > 0
    )


def build_rollup(budget_line: str, people: list[Person]) -> BudgetLineRollup:
    """Sum a slice of people into one BudgetLineRollup."""
    amended = sum(p.amended_budget_total for p in people)
    reference = AMENDMENT_1_TOTALS.get(budget_line)
    delta = (amended - reference) if reference is not None else None
    reconciles = (
        abs(delta) <= RECONCILIATION_TOLERANCE_USD
        if delta is not None else None
    )
    return BudgetLineRollup(
        budget_line=budget_line,
        label=BUDGET_LINE_LABELS.get(budget_line, budget_line),
        person_count=len(people),
        amended_budget_total=amended,
        paid_to_date=sum(p.paid_to_date for p in people),
        projected_total_remaining=sum(p.projected_total_remaining for p in people),
        total_committed=sum(p.total_committed for p in people),
        variance_vs_amended=sum(p.variance_vs_amended for p in people),
        amendment_1_reference=reference,
        reconciliation_delta=delta,
        reconciles=reconciles,
    )


def check_reconciliation(rollups: Iterable[BudgetLineRollup]) -> list[ReconciliationWarning]:
    """Generate a warning per rollup that fails Amendment 1 reconciliation.

    Hard constraint per the spec: silent reconciliation drift is exactly
    the failure mode the view is designed to prevent.
    """
    issues: list[ReconciliationWarning] = []
    for r in rollups:
        if r.reconciles is False:
            issues.append(ReconciliationWarning(
                level="error",
                budget_line=r.budget_line,
                message=(
                    f"{r.label}: amended budget total ${r.amended_budget_total:,.2f} "
                    f"does NOT reconcile to Amendment 1 reference "
                    f"${r.amendment_1_reference:,.2f} "
                    f"(delta ${r.reconciliation_delta:+,.2f})."
                ),
            ))
        elif r.reconciles is None and r.amended_budget_total > 0:
            issues.append(ReconciliationWarning(
                level="info",
                budget_line=r.budget_line,
                message=(
                    f"{r.label}: no Amendment 1 reference recorded for "
                    f"this budget line; rollup ${r.amended_budget_total:,.2f} "
                    f"is unverified."
                ),
            ))
    return issues


# ---------------------------------------------------------------------------
# Public extract entrypoint
# ---------------------------------------------------------------------------

def extract_personnel_and_contractors(
    workbook_path: str | Path | None = None,
    *,
    write_snapshot: bool = True,
) -> PersonnelExtract:
    """Read the personnel workbook, compute derived fields, return extract.

    Parameters
    ----------
    workbook_path
        Path to `K8341_Personnel_and_Contractors.xlsx`. If None, defaults
        to `agents/finance/design/fixtures/K8341_Personnel_and_Contractors.xlsx`
        (same fixtures dir cockpit_data.py uses).
    write_snapshot
        If True (default), writes a sibling JSON snapshot for diffability
        in PRs. Set False in tests / read-only environments.
    """
    if workbook_path is None:
        # Match the default fixtures-dir convention from cockpit_data.py.
        # Repo root = three parents up from this file
        # (agents/finance/personnel.py → wfd-os/).
        repo_root = Path(__file__).resolve().parents[2]
        workbook_path = (
            repo_root / "agents" / "finance" / "design" / "fixtures"
            / "K8341_Personnel_and_Contractors.xlsx"
        )
    workbook_path = Path(workbook_path)

    people, parse_warnings = parse_workbook(workbook_path)

    for person in people:
        compute_derived(person)

    # Group by budget line; one rollup per known line plus any
    # unrecognized line that actually has people in it.
    by_line: dict[str, list[Person]] = {bl: [] for bl in VALID_BUDGET_LINES}
    for person in people:
        by_line.setdefault(person.budget_line, []).append(person)
    rollups = [
        build_rollup(line, by_line.get(line, []))
        for line in (
            list(BUDGET_LINE_LABELS.keys())
            + sorted(set(by_line) - VALID_BUDGET_LINES)
        )
        if line  # skip empty-string line (people whose budget_line was blank)
    ]

    reconciliation_issues = check_reconciliation(rollups)
    all_warnings = parse_warnings + reconciliation_issues

    distinct_count = len({p.name for p in people if p.name})
    summary_paid = sum(p.paid_to_date for p in people)
    summary_committed = sum(p.total_committed for p in people)
    summary_variance = sum(p.variance_vs_amended for p in people)

    extract = PersonnelExtract(
        people=people,
        rollups=rollups,
        distinct_person_count=distinct_count,
        summary_paid_to_date=summary_paid,
        summary_total_committed=summary_committed,
        summary_variance_vs_amended=summary_variance,
        reconciliation_warnings=all_warnings,
        extracted_at=datetime.utcnow().isoformat() + "Z",
        source_workbook=str(workbook_path) if workbook_path.exists() else None,
    )

    if write_snapshot and workbook_path.exists():
        _write_snapshot(workbook_path, extract)

    return extract


def _write_snapshot(workbook_path: Path, extract: PersonnelExtract) -> None:
    """Write a JSON snapshot of the extract sibling to the workbook."""
    snap_path = workbook_path.with_name(workbook_path.stem + ".snapshot.json")
    payload = to_dict(extract)
    snap_path.write_text(
        json.dumps(payload, indent=2, default=str),
        encoding="utf-8",
    )


# ---------------------------------------------------------------------------
# Serialization — for cockpit_api.py JSON responses + snapshot
# ---------------------------------------------------------------------------

def to_dict(extract: PersonnelExtract) -> dict:
    """Serialize a PersonnelExtract to a JSON-friendly dict.

    Used by:
      - cockpit_api.py to respond to /cockpit/tabs/budget (personnel field)
      - the snapshot file written sibling to the workbook
      - the HTML template via Jinja (passed through .__dict__-style access)
    """
    def person_to_dict(p: Person) -> dict:
        d = asdict(p)
        # Add derived/property fields not on the dataclass itself.
        d["documentation_incomplete"] = p.documentation_incomplete
        d["drill_key"] = p.drill_key
        return d

    return {
        "people": [person_to_dict(p) for p in extract.people],
        "rollups": [asdict(r) for r in extract.rollups],
        "distinct_person_count": extract.distinct_person_count,
        "summary": {
            "paid_to_date": extract.summary_paid_to_date,
            "total_committed": extract.summary_total_committed,
            "variance_vs_amended": extract.summary_variance_vs_amended,
        },
        "reconciliation_warnings": [asdict(w) for w in extract.reconciliation_warnings],
        "extracted_at": extract.extracted_at,
        "source_workbook": extract.source_workbook,
    }


# ---------------------------------------------------------------------------
# Drill builder — emit `person:<id>` drills for the cockpit drill registry
# ---------------------------------------------------------------------------

def build_personnel_drills(extract: PersonnelExtract) -> dict:
    """Return a dict of drill entries, keyed `person:<id>`.

    Shape matches the cockpit's polymorphic drill schema in
    agents/finance/design/cockpit_data.py — section ids are stable
    (`identity`, `rate_detail`, `actuals`, `projections`,
    `qb_reconciliation`, `documentation_status`).
    """
    drills: dict[str, dict] = {}

    for p in extract.people:
        sections: list[dict] = []

        # Identity ----------------------------------------------------------
        identity_rows = [
            {"label": "Name", "value": p.name or "—"},
            {"label": "Role", "value": p.role or "—"},
            {"label": "Engagement type", "value": p.engagement_type or "—"},
            {"label": "Budget line",
             "value": BUDGET_LINE_LABELS.get(p.budget_line, p.budget_line or "—")},
            {"label": "Vendor legal entity",
             "value": p.vendor_legal_entity or "—"},
            {"label": "Start date",
             "value": p.start_date.isoformat() if p.start_date else "—"},
            {"label": "End date",
             "value": p.end_date.isoformat() if p.end_date else "Through grant end"},
        ]
        sections.append({
            "id": "identity",
            "type": "rows",
            "title": "Identity",
            "rows": identity_rows,
        })

        # Rate detail -------------------------------------------------------
        rate_rows = [
            {"label": "Rate",
             "value": f"${p.rate_amount:,.2f} / {p.rate_unit}"
             if p.rate_amount and p.rate_unit else "—"},
            {"label": "Rate effective date",
             "value": p.rate_effective_date.isoformat()
             if p.rate_effective_date else "—"},
            {"label": "Rate basis",
             "value": p.rate_basis or "Not documented"},
            {"label": "Amended budget total",
             "value": f"${p.amended_budget_total:,.2f}"},
        ]
        sections.append({
            "id": "rate_detail",
            "type": "rows",
            "title": "Rate detail",
            "rows": rate_rows,
        })

        # Quarterly actuals -------------------------------------------------
        if p.actuals:
            sections.append({
                "id": "actuals",
                "type": "table",
                "title": "Quarterly actuals",
                "columns": [
                    {"key": "quarter", "label": "Quarter"},
                    {"key": "amount_paid", "label": "Amount paid",
                     "align": "right", "numeric": True},
                    {"key": "source", "label": "Source"},
                    {"key": "qb_vendor_name", "label": "QB vendor name"},
                ],
                "rows": [
                    {
                        "quarter": a.quarter,
                        "amount_paid": f"${a.amount_paid:,.2f}",
                        "source": a.source,
                        "qb_vendor_name": a.qb_vendor_name or "—",
                    }
                    for a in sorted(p.actuals, key=lambda x: x.quarter)
                ],
                "note": f"Paid to date: ${p.paid_to_date:,.2f}",
            })
        else:
            sections.append({
                "id": "actuals",
                "type": "prose",
                "title": "Quarterly actuals",
                "body": "No actuals recorded yet for this person.",
            })

        # Quarterly projections --------------------------------------------
        if p.projections:
            sections.append({
                "id": "projections",
                "type": "table",
                "title": "Quarterly projections",
                "columns": [
                    {"key": "quarter", "label": "Quarter"},
                    {"key": "projected_amount", "label": "Projected",
                     "align": "right", "numeric": True},
                    {"key": "projection_basis", "label": "Basis"},
                ],
                "rows": [
                    {
                        "quarter": pr.quarter,
                        "projected_amount": f"${pr.projected_amount:,.2f}",
                        "projection_basis": pr.projection_basis or "—",
                    }
                    for pr in sorted(p.projections, key=lambda x: x.quarter)
                ],
                "note": (
                    f"Projected remaining: ${p.projected_total_remaining:,.2f} · "
                    f"Total committed: ${p.total_committed:,.2f} · "
                    f"Variance vs amended: ${p.variance_vs_amended:+,.2f}"
                ),
            })
        else:
            sections.append({
                "id": "projections",
                "type": "prose",
                "title": "Quarterly projections",
                "body": (
                    "Projections not yet recorded for remaining quarters. "
                    "Per spec: missing projections are flagged, not "
                    "defaulted to zero."
                    if p.projections_missing
                    else "No remaining quarters to project (person concluded)."
                ),
            })

        # QB reconciliation -------------------------------------------------
        qb_vendors = sorted({
            a.qb_vendor_name for a in p.actuals if a.qb_vendor_name
        })
        sections.append({
            "id": "qb_reconciliation",
            "type": "rows",
            "title": "QB reconciliation",
            "rows": [
                {"label": "Display name", "value": p.name or "—"},
                {"label": "QB vendor name(s)",
                 "value": ", ".join(qb_vendors) if qb_vendors else "—"},
                {"label": "Paid-to-date (sum of actuals)",
                 "value": f"${p.paid_to_date:,.2f}"},
            ],
            "note": (
                "Compare against QB vendor history. v1: manual reconciliation. "
                "Automated QB sync deferred to v1.3.0 wfdos-common per spec."
            ),
        })

        # Documentation status (honesty section) ----------------------------
        if p.documentation_incomplete:
            sections.append({
                "id": "documentation_status",
                "type": "rows",
                "title": "Documentation status",
                "rows": [
                    {"label": "Missing required field",
                     "value": _humanize_field(field_name)}
                    for field_name in p.missing_required_fields
                ],
                "note": (
                    "Per spec: rows missing required fields are flagged "
                    "rather than silently filled. Update the workbook to "
                    "clear these flags."
                ),
            })

        # Status chip -------------------------------------------------------
        chip = None
        if p.documentation_incomplete:
            chip = {"label": "Documentation incomplete", "tone": "watch"}
        elif p.variance_vs_amended < 0:
            chip = {"label": "Projected overrun", "tone": "critical"}

        entry = {
            "eyebrow": (
                "Person" if p.engagement_type == "employee"
                else "Contractor" if p.engagement_type == "contractor"
                else "Subcontractor" if p.engagement_type == "subcontractor"
                else "Personnel"
            ),
            "title": p.name or p.id,
            "summary": (
                f"{p.role or 'Role TBD'} · "
                f"{BUDGET_LINE_LABELS.get(p.budget_line, p.budget_line or 'budget line TBD')} · "
                f"variance ${p.variance_vs_amended:+,.0f}"
            ),
            "sections": sections,
        }
        if chip:
            entry["status_chip"] = chip

        drills[p.drill_key] = entry

    return drills


def _humanize_field(field_name: str) -> str:
    """`rate_basis` → `Rate basis` for display in the documentation_status section."""
    return field_name.replace("_", " ").capitalize()
