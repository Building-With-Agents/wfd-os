"""Generate a minimal 2-3 contract sample workbook for engine-side testing.

Produces K8341_Contracts_sample.xlsx with three contracts (one per
contract_type that the reconciliation cares about), one amendment, and
one termination — exercising every loader code path. Values are
NOT intended to reconcile to Amendment 1; the sample purposely shows
drift so reconciliation warnings appear during smoke-testing.

Run from the agents/grant-compliance directory:
    python data/contracts_bootstrap/generate_sample.py

This script ships in the repo so anyone can regenerate the sample.
Krista's full data is a separate Excel file populated against the same
template — it lives outside the repo.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook


SAMPLE_PATH = Path(__file__).parent / "K8341_Contracts_sample.xlsx"


def build_workbook() -> Workbook:
    wb = Workbook()
    # Drop the default sheet — we name them explicitly.
    default = wb.active
    if default is not None:
        wb.remove(default)

    # ----- Contracts sheet -----
    ws = wb.create_sheet("Contracts")
    ws.append([
        "record_key",
        "vendor_name_display",
        "vendor_legal_entity",
        "vendor_qb_names",
        "contract_type",
        "compliance_classification",
        "classification_rationale",
        "procurement_method",
        "original_executed_date",
        "original_effective_date",
        "current_end_date",
        "original_contract_value_dollars",
        "current_contract_value_dollars",
        "status",
        "payment_basis",
        "payment_basis_detail",
        "executed_contract_link",
        "scope_of_work_summary",
        "notes",
    ])
    # Training provider — sample value, intentional drift vs $2.3M Amendment 1
    ws.append([
        "ada",
        "Ada Developers Academy",
        "Ada Developers Academy",
        "Ada|Ada Developers Academy|ADA",
        "training_provider",
        "contractor_200_331b",
        "Performance criteria + scope tightly defined; non-Federal compliance burden.",
        "competitive_proposals",
        date(2024, 7, 1),
        date(2024, 7, 1),
        date(2026, 6, 30),
        500_000.00,
        500_000.00,
        "active",
        "per_placement",
        "$5,000 per confirmed placement, capped at 100 placements.",
        "https://example.invalid/contracts/ada-2024.pdf",
        "Three-cohort apprenticeship-prep program for software engineering roles.",
        "Sample data — not reconciled.",
    ])
    # Strategic partner subrecipient
    ws.append([
        "wabs",
        "WABS",
        "Washington Alliance for Better Schools",
        "WABS|Washington Alliance for Better Schools",
        "strategic_partner_subrecipient",
        "subrecipient_200_331a",
        "Programmatic decisions made; Federal compliance pass-through.",
        "sole_source",
        date(2024, 9, 1),
        date(2024, 9, 1),
        date(2025, 12, 31),
        850_000.00,
        850_000.00,
        "terminated_by_funder",
        "milestone",
        "Milestone-based payments tied to enrollment + completion.",
        "https://example.invalid/contracts/wabs-2024.pdf",
        "Subaward to WABS for K-12 districts integration of WFD pipeline.",
        "Sample termination flow exercised below.",
    ])
    # CFA contractor — intentional drift vs $1.02M Amendment 1
    ws.append([
        "ai_engage",
        "AI Engage",
        "AI Engage Group LLC",
        "AI Engage|AIE|Jason Mangold|AI Engage Group LLC",
        "cfa_contractor",
        "contractor_200_331b",
        "Service vendor; CFA controls scope and acceptance.",
        "competitive_rfp",
        date(2025, 1, 15),
        date(2025, 2, 1),
        date(2026, 8, 31),
        300_000.00,
        375_000.00,
        "active",
        "fixed_fee",
        "Three milestone payments aligned to the AIE delivery plan.",
        "https://example.invalid/contracts/aie-2025.pdf",
        "Ongoing managed-service support for CFA platform infrastructure.",
        "Sample data — current value reflects amendment 1 below.",
    ])

    # ----- Amendments sheet -----
    ws_a = wb.create_sheet("Amendments")
    ws_a.append([
        "contract_record_key",
        "amendment_number",
        "amendment_type",
        "executed_date",
        "effective_date",
        "previous_value_dollars",
        "new_value_dollars",
        "previous_end_date",
        "new_end_date",
        "summary_of_changes",
        "document_link",
    ])
    ws_a.append([
        "ai_engage",
        1,
        "value_change",
        date(2025, 6, 1),
        date(2025, 6, 1),
        300_000.00,
        375_000.00,
        date(2026, 1, 31),
        date(2026, 8, 31),
        "Scope expansion + period extension to support Q3 platform work.",
        "https://example.invalid/contracts/aie-2025-amendment1.pdf",
    ])

    # ----- Terminations sheet -----
    ws_t = wb.create_sheet("Terminations")
    ws_t.append([
        "contract_record_key",
        "terminated_by",
        "termination_date",
        "termination_reason",
        "termination_correspondence_link",
        "final_reconciliation_link",
        "closeout_findings",
    ])
    ws_t.append([
        "wabs",
        "funder",
        date(2025, 12, 31),
        "ESD instructed CFA to wind down WABS subaward as part of grant restructure.",
        "https://example.invalid/correspondence/wabs-termination.pdf",
        "https://example.invalid/contracts/wabs-final-recon.pdf",
        "No findings; subaward closed cleanly.",
    ])

    return wb


if __name__ == "__main__":
    wb = build_workbook()
    wb.save(SAMPLE_PATH)
    print(f"Wrote {SAMPLE_PATH}")
