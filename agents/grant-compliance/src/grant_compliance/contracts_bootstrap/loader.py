"""Excel loader for the contracts inventory bootstrap.

Reads the K8341_Contracts.xlsx workbook (three sheets — Contracts,
Amendments, Terminations) and produces validated Pydantic records.

Contract: the loader trusts only what's in the file. It does NOT
infer missing fields, fall back to defaults beyond what's in
schemas.py, or accept unknown columns. Schema drift surfaces as
LoaderError rather than silently parsing a partial record.

Money columns in Excel arrive in dollars (e.g. 250000.00 for $250K)
because that's what humans type. The loader converts to cents for
the database.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]

from grant_compliance.contracts_bootstrap.schemas import (
    AmendmentRecord,
    ContractRecord,
    LoadedBundle,
    TerminationRecord,
)


class LoaderError(ValueError):
    """Raised when the workbook structure is invalid (missing sheet,
    missing required column, unparseable value). Carries a human-
    readable message intended for the CLI to display verbatim.
    """


# Required column names per sheet. Columns may appear in any order in
# the spreadsheet (we look up by header name), but every required column
# must be present. Optional columns may be omitted.
_CONTRACTS_REQUIRED = (
    "record_key",
    "vendor_name_display",
    "vendor_legal_entity",
    "contract_type",
    "original_contract_value_dollars",
    "current_contract_value_dollars",
)
_CONTRACTS_OPTIONAL = (
    "vendor_qb_names",
    "compliance_classification",
    "classification_rationale",
    "procurement_method",
    "original_executed_date",
    "original_effective_date",
    "current_end_date",
    "status",
    "payment_basis",
    "payment_basis_detail",
    "executed_contract_link",
    "scope_of_work_summary",
    "notes",
)

_AMENDMENTS_REQUIRED = (
    "contract_record_key",
    "amendment_number",
    "amendment_type",
)
_AMENDMENTS_OPTIONAL = (
    "executed_date",
    "effective_date",
    "previous_value_dollars",
    "new_value_dollars",
    "previous_end_date",
    "new_end_date",
    "summary_of_changes",
    "document_link",
)

_TERMINATIONS_REQUIRED = (
    "contract_record_key",
    "terminated_by",
)
_TERMINATIONS_OPTIONAL = (
    "termination_date",
    "termination_reason",
    "termination_correspondence_link",
    "final_reconciliation_link",
    "closeout_findings",
)


def load_workbook_bundle(path: str | Path) -> LoadedBundle:
    """Read an Excel workbook and return a LoadedBundle.

    Raises LoaderError on any structural problem. Caller (typically the
    CLI) should catch and display the message.
    """
    p = Path(path)
    if not p.exists():
        raise LoaderError(f"File not found: {p}")

    try:
        wb = load_workbook(p, data_only=True, read_only=True)
    except Exception as exc:
        raise LoaderError(f"Could not open workbook: {exc}") from exc

    sheet_names = set(wb.sheetnames)
    for required in ("Contracts",):
        if required not in sheet_names:
            raise LoaderError(
                f"Workbook missing required sheet '{required}'. "
                f"Found: {sorted(sheet_names)}"
            )

    contracts = _load_contracts_sheet(wb["Contracts"])

    amendments: list[AmendmentRecord] = []
    if "Amendments" in sheet_names:
        amendments = _load_amendments_sheet(wb["Amendments"])

    terminations: list[TerminationRecord] = []
    if "Terminations" in sheet_names:
        terminations = _load_terminations_sheet(wb["Terminations"])

    bundle = LoadedBundle(
        contracts=contracts,
        amendments=amendments,
        terminations=terminations,
    )

    cross_ref_errors = bundle.validate_cross_references()
    if cross_ref_errors:
        raise LoaderError(
            "Cross-reference errors in workbook:\n  - "
            + "\n  - ".join(cross_ref_errors)
        )

    return bundle


# ---------- per-sheet readers ----------


def _load_contracts_sheet(ws: Any) -> list[ContractRecord]:
    rows = _read_sheet_rows(ws, _CONTRACTS_REQUIRED, _CONTRACTS_OPTIONAL, "Contracts")
    out: list[ContractRecord] = []
    for row_num, row in rows:
        try:
            record = ContractRecord(
                record_key=_required_str(row, "record_key", row_num, "Contracts"),
                vendor_name_display=_required_str(
                    row, "vendor_name_display", row_num, "Contracts"
                ),
                vendor_legal_entity=_required_str(
                    row, "vendor_legal_entity", row_num, "Contracts"
                ),
                vendor_qb_names=row.get("vendor_qb_names"),
                contract_type=_required_str(row, "contract_type", row_num, "Contracts"),
                compliance_classification=row.get("compliance_classification")
                or "unclassified",
                classification_rationale=row.get("classification_rationale"),
                procurement_method=row.get("procurement_method") or "unknown",
                original_executed_date=_to_date(row.get("original_executed_date")),
                original_effective_date=_to_date(row.get("original_effective_date")),
                current_end_date=_to_date(row.get("current_end_date")),
                original_contract_value_cents=_dollars_to_cents(
                    _required_number(
                        row, "original_contract_value_dollars", row_num, "Contracts"
                    )
                ),
                current_contract_value_cents=_dollars_to_cents(
                    _required_number(
                        row, "current_contract_value_dollars", row_num, "Contracts"
                    )
                ),
                status=row.get("status") or "active",
                payment_basis=row.get("payment_basis") or "other",
                payment_basis_detail=row.get("payment_basis_detail"),
                executed_contract_link=row.get("executed_contract_link"),
                scope_of_work_summary=row.get("scope_of_work_summary"),
                notes=row.get("notes"),
            )
        except (ValueError, TypeError) as exc:
            raise LoaderError(
                f"Contracts row {row_num}: {exc}"
            ) from exc
        out.append(record)
    return out


def _load_amendments_sheet(ws: Any) -> list[AmendmentRecord]:
    rows = _read_sheet_rows(
        ws, _AMENDMENTS_REQUIRED, _AMENDMENTS_OPTIONAL, "Amendments"
    )
    out: list[AmendmentRecord] = []
    for row_num, row in rows:
        try:
            record = AmendmentRecord(
                contract_record_key=_required_str(
                    row, "contract_record_key", row_num, "Amendments"
                ),
                amendment_number=int(
                    _required_number(row, "amendment_number", row_num, "Amendments")
                ),
                amendment_type=_required_str(
                    row, "amendment_type", row_num, "Amendments"
                ),
                executed_date=_to_date(row.get("executed_date")),
                effective_date=_to_date(row.get("effective_date")),
                previous_value_cents=_optional_dollars_to_cents(
                    row.get("previous_value_dollars")
                ),
                new_value_cents=_optional_dollars_to_cents(
                    row.get("new_value_dollars")
                ),
                previous_end_date=_to_date(row.get("previous_end_date")),
                new_end_date=_to_date(row.get("new_end_date")),
                summary_of_changes=row.get("summary_of_changes"),
                document_link=row.get("document_link"),
            )
        except (ValueError, TypeError) as exc:
            raise LoaderError(
                f"Amendments row {row_num}: {exc}"
            ) from exc
        out.append(record)
    return out


def _load_terminations_sheet(ws: Any) -> list[TerminationRecord]:
    rows = _read_sheet_rows(
        ws, _TERMINATIONS_REQUIRED, _TERMINATIONS_OPTIONAL, "Terminations"
    )
    out: list[TerminationRecord] = []
    for row_num, row in rows:
        try:
            record = TerminationRecord(
                contract_record_key=_required_str(
                    row, "contract_record_key", row_num, "Terminations"
                ),
                terminated_by=_required_str(
                    row, "terminated_by", row_num, "Terminations"
                ),
                termination_date=_to_date(row.get("termination_date")),
                termination_reason=row.get("termination_reason"),
                termination_correspondence_link=row.get(
                    "termination_correspondence_link"
                ),
                final_reconciliation_link=row.get("final_reconciliation_link"),
                closeout_findings=row.get("closeout_findings"),
            )
        except (ValueError, TypeError) as exc:
            raise LoaderError(
                f"Terminations row {row_num}: {exc}"
            ) from exc
        out.append(record)
    return out


# ---------- worksheet → row dicts ----------


def _read_sheet_rows(
    ws: Any,
    required: tuple[str, ...],
    optional: tuple[str, ...],
    sheet_name: str,
) -> list[tuple[int, dict[str, Any]]]:
    """Read a worksheet into a list of (excel_row_number, row_dict) pairs.

    Header row is row 1. Data starts row 2. Empty rows (every cell None
    or blank) are skipped silently — Excel often leaves trailing empty
    rows after the data ends.
    """
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration as exc:
        raise LoaderError(f"Sheet '{sheet_name}' is empty.") from exc

    header_clean = [str(h).strip() if h is not None else None for h in header]
    name_to_idx = {h: i for i, h in enumerate(header_clean) if h is not None}

    missing = [r for r in required if r not in name_to_idx]
    if missing:
        raise LoaderError(
            f"Sheet '{sheet_name}' missing required columns: {missing}. "
            f"Found columns: {[h for h in header_clean if h]}"
        )

    unknown = [
        h
        for h in header_clean
        if h is not None and h not in required and h not in optional
    ]
    if unknown:
        raise LoaderError(
            f"Sheet '{sheet_name}' has unknown columns: {unknown}. "
            f"Allowed: {sorted(set(required) | set(optional))}"
        )

    out: list[tuple[int, dict[str, Any]]] = []
    for excel_row_num, row in enumerate(rows_iter, start=2):
        if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row):
            continue
        row_dict: dict[str, Any] = {}
        for col_name, idx in name_to_idx.items():
            if idx < len(row):
                v = row[idx]
                if isinstance(v, str):
                    v = v.strip() or None
                row_dict[col_name] = v
        out.append((excel_row_num, row_dict))
    return out


# ---------- value coercion ----------


def _required_str(row: dict[str, Any], key: str, row_num: int, sheet: str) -> str:
    v = row.get(key)
    if v is None or (isinstance(v, str) and not v.strip()):
        raise LoaderError(
            f"Sheet '{sheet}' row {row_num}: required column '{key}' is empty"
        )
    return str(v).strip()


def _required_number(
    row: dict[str, Any], key: str, row_num: int, sheet: str
) -> float:
    v = row.get(key)
    if v is None:
        raise LoaderError(
            f"Sheet '{sheet}' row {row_num}: required column '{key}' is empty"
        )
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").replace("$", "").strip())
    except ValueError as exc:
        raise LoaderError(
            f"Sheet '{sheet}' row {row_num}: column '{key}' is not numeric: {v!r}"
        ) from exc


def _to_date(v: Any) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        # Accept ISO format only — reject ambiguous regional formats so
        # 03/04/2025 doesn't silently get parsed as Mar 4 vs Apr 3.
        try:
            return date.fromisoformat(s)
        except ValueError as exc:
            raise LoaderError(
                f"Date value {v!r} is not ISO format (YYYY-MM-DD). "
                f"Excel typed dates are accepted directly; for text cells "
                f"use ISO format only."
            ) from exc
    raise LoaderError(f"Cannot interpret value {v!r} as a date.")


def _dollars_to_cents(dollars: float) -> int:
    """Convert dollars (Excel-typical) to cents (engine convention).

    Uses banker's rounding to nearest cent — should be a no-op for
    cleanly typed dollar amounts but tolerates floating-point artifacts
    like 250000.0000001.
    """
    return round(dollars * 100)


def _optional_dollars_to_cents(dollars: Any) -> int | None:
    if dollars is None or (isinstance(dollars, str) and not dollars.strip()):
        return None
    if isinstance(dollars, (int, float)):
        return _dollars_to_cents(float(dollars))
    try:
        return _dollars_to_cents(
            float(str(dollars).replace(",", "").replace("$", "").strip())
        )
    except ValueError as exc:
        raise LoaderError(f"Value {dollars!r} is not numeric.") from exc
