import io
import openpyxl
import pandas as pd
from database.db import get_session
from database.models import BaselineData, MonthlySnapshot


# Keywords used to match actual SharePoint filenames to data types.
# Actual filenames may have spaces, suffixes like (1), etc.
BASELINE_TYPE_KEYWORDS = {
    "provider_reconciliation": ["provider_reconciliation", "provider reconciliation"],
    "contractors": ["gjc_contractors", "gjc contractors"],
    "outcomes": ["partner_data_outcomes", "partner data outcomes"],
    "cost_per_placement": ["cost_per_placement", "cost per placement"],
    "budget_exhibit": ["exh b", "exh_b", "exhibit"],
}


def _classify_file(filename: str) -> str:
    """Match a filename to a data type using keyword matching."""
    name_lower = filename.lower()
    for data_type, keywords in BASELINE_TYPE_KEYWORDS.items():
        if any(kw in name_lower for kw in keywords):
            return data_type
    return "unknown"


def _parse_excel_to_rows(data: bytes, filename: str) -> list[dict]:
    """Parse an Excel file into a list of row dicts, one per row per sheet."""
    xls = pd.ExcelFile(io.BytesIO(data), engine="openpyxl")
    rows = []
    for sheet in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet, engine="openpyxl")
        df = df.dropna(how="all")
        df.columns = df.columns.astype(str).str.strip()
        for _, row in df.iterrows():
            row_dict = {}
            for col, val in row.items():
                if pd.isna(val):
                    row_dict[col] = None
                elif hasattr(val, "isoformat"):
                    row_dict[col] = val.isoformat()
                else:
                    row_dict[col] = val
            rows.append({"sheet_name": sheet, "row_data": row_dict})
    return rows


def _parse_partner_outcomes(data: bytes, filename: str) -> list[dict]:
    """
    Targeted parser for Partner Data Outcomes Summary.xlsx.
    Reads ONLY the 'Updated Ext Actual New Contract' sheet.
    Extracts Job Placements rows using exact column positions:
      Col A (0): Provider name (appears in header rows above each section)
      Col C (2): Metric name — we only keep rows where this is "Job Placements"
      Actual placement counts are in the even-indexed columns:
        E(4)=Q1 24, G(6)=Q2 24, I(8)=Q3 24, K(10)=Q4 24,
        M(12)=Q1 25, O(14)=Q2 25, Q(16)=Q3 25, S(18)=Q4 25
    """
    SHEET = "Updated Ext Actual New Contract"
    QUARTERS = ["Q1_2024", "Q2_2024", "Q3_2024", "Q4_2024",
                "Q1_2025", "Q2_2025", "Q3_2025", "Q4_2025"]
    ACTUAL_COLS = [4, 6, 8, 10, 12, 14, 16, 18]  # E, G, I, K, M, O, Q, S

    df = pd.read_excel(io.BytesIO(data), sheet_name=SHEET,
                       engine="openpyxl", header=None)

    rows = []
    current_provider = None
    grand_total = 0

    for idx, row in df.iterrows():
        col_a = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
        col_c = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ""

        # Provider names appear in col A on header rows (col C is empty)
        if col_a and col_c == "" and col_a not in ("NaN",) \
                and "NOTE" not in col_a and "Quarter" not in col_a \
                and not col_a.startswith("Q"):
            current_provider = col_a.replace("\n", " ").strip()

        if col_c != "Job Placements" or not current_provider:
            continue

        # Read actual placement counts from the fixed columns
        quarterly = {}
        provider_total = 0
        for q, ci in zip(QUARTERS, ACTUAL_COLS):
            val = row.iloc[ci] if ci < len(row) and pd.notna(row.iloc[ci]) else 0
            count = int(val) if isinstance(val, (int, float)) else 0
            quarterly[q] = count
            provider_total += count

        rows.append({
            "sheet_name": "provider_placements",
            "row_data": {
                "provider": current_provider,
                "quarterly_actuals": quarterly,
                "total_placements": provider_total,
            }
        })
        grand_total += provider_total

    # Add the all-provider summary with the confirmed net total
    rows.append({
        "sheet_name": "placement_summary",
        "row_data": {
            "metric": "Net Job Placements through Q4 2025",
            "spreadsheet_gross_total": grand_total,
            "net_placements": 423,
            "retractions": grand_total - 423,
            "pip_threshold": 730,
            "pct_of_pip": round(423 / 730 * 100, 1),
            "placements_remaining": 307,
            "note": f"Gross placements from spreadsheet: {grand_total}. "
                    f"After {grand_total - 423} retractions, net is 423.",
        }
    })

    return rows


def _parse_provider_reconciliation(data: bytes, filename: str) -> list[dict]:
    """
    Targeted parser for K8341_Provider_Reconciliation_v3_3-27.xlsx.
    Reads 'Provider Reconciliation' sheet using openpyxl.
    Structure:
      Col A: Provider name
      Col B: Amended Budget (Exh B Amend 1)
      Col C: Tracker Total
      Col D: QB Actual (as of 3/26/26)
      Col E: Balance Remaining
      Col F: Notes / Flags
    """
    from openpyxl.utils import get_column_letter
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb["Provider Reconciliation"]

    rows = []
    for row_idx in range(5, ws.max_row + 1):
        col_a = ws.cell(row=row_idx, column=1).value  # Provider
        col_b = ws.cell(row=row_idx, column=2).value  # Budget
        col_d = ws.cell(row=row_idx, column=4).value  # QB Actual
        col_e = ws.cell(row=row_idx, column=5).value  # Balance
        col_f = ws.cell(row=row_idx, column=6).value  # Notes

        if not col_a:
            continue

        name = str(col_a).strip()

        # Skip section headers and subtotal/total rows
        if name in ("ESD-DIRECTED CLAWBACKS — 3 PROVIDERS",
                     "ACTIVE TRAINING PROVIDERS — 6 CONTRACTS (Q1 placement data due April 6)",
                     "CLOSED PROVIDERS — DELIVERED PLACEMENTS",
                     "CLOSED PROVIDERS — SUPPORT / ENGAGEMENT (not placement-based)",
                     "CFA CONTRACTORS — incl AI Engage  (QB actuals pending Krista)",
                     "CFA SALARIES & BENEFITS  (QB actuals pending Krista)",
                     "ITEMS REQUIRING IMMEDIATE CLARIFICATION WITH KRISTA",
                     "FULL GRANT TOTAL (Amendment 1)",
                     "#"):
            continue

        # Parse QB actual — may be a number, "Pending QB", "Not in QB", etc.
        qb_actual = None
        if isinstance(col_d, (int, float)):
            qb_actual = float(col_d)

        budget = float(col_b) if isinstance(col_b, (int, float)) else None
        balance = float(col_e) if isinstance(col_e, (int, float)) else None

        rows.append({
            "sheet_name": "provider_reconciliation",
            "row_data": {
                "provider": name,
                "amended_budget": budget,
                "qb_actual": qb_actual,
                "balance_remaining": balance,
                "notes": str(col_f) if col_f else None,
            }
        })

    return rows


def _seed_budget_snapshot():
    """
    Seed a MonthlySnapshot with confirmed QB actuals as of 3/26/2026.
    These are hardcoded from the verified Provider Reconciliation spreadsheet.

    Confirmed totals:
      GJC Contractors (all training + support providers): $1,328,794
      CFA Contractors (AI Engage + Vargo):                $  810,607
      Salaries + Benefits + Other + Indirect (lump):      $1,227,851
      TOTAL:                                              $3,367,252

    The salary/benefits/other/indirect breakdown is pending Krista,
    so we prorate the $1,227,851 lump sum by budget ratios.
    """
    # Confirmed QB actuals
    gjc_spent = 1328794.00
    cfa_spent = 810607.00
    overhead_lump = 1227851.00

    # Prorate overhead lump across 4 categories by budget ratio
    salary_budget = 1097662.41
    benefits_budget = 173169.94
    other_budget = 88921.06
    indirect_budget = 178798.91
    overhead_total = salary_budget + benefits_budget + other_budget + indirect_budget

    salary_spent = round(overhead_lump * (salary_budget / overhead_total), 2)
    benefits_spent = round(overhead_lump * (benefits_budget / overhead_total), 2)
    other_spent = round(overhead_lump * (other_budget / overhead_total), 2)
    indirect_spent = round(overhead_lump * (indirect_budget / overhead_total), 2)

    session = get_session()
    try:
        session.query(MonthlySnapshot).filter(MonthlySnapshot.month == "2026-03").delete()

        snapshot = MonthlySnapshot(
            month="2026-03",
            gjc_contractors_spent=gjc_spent,
            cfa_contractors_spent=cfa_spent,
            personnel_salaries_spent=salary_spent,
            personnel_benefits_spent=benefits_spent,
            other_direct_spent=other_spent,
            indirect_costs_spent=indirect_spent,
            raw_data={
                "source": "Confirmed QB actuals as of 2026-03-26",
                "gjc_contractors": gjc_spent,
                "cfa_contractors": cfa_spent,
                "overhead_lump_sum": overhead_lump,
                "overhead_note": "Salary/benefits/other/indirect prorated by budget ratio. Individual breakdown pending.",
                "total": gjc_spent + cfa_spent + overhead_lump,
            },
        )
        session.add(snapshot)
        session.commit()
        return {
            "gjc_contractors": gjc_spent,
            "cfa_contractors": cfa_spent,
            "personnel_salaries": salary_spent,
            "personnel_benefits": benefits_spent,
            "other_direct": other_spent,
            "indirect_costs": indirect_spent,
            "total": gjc_spent + cfa_spent + overhead_lump,
        }
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


async def load_baseline_from_sharepoint():
    """Download baseline files from SharePoint and seed the database."""
    from ingestion.sharepoint import get_baseline_file_set

    files = await get_baseline_file_set()
    if not files:
        return {"status": "error", "message": "No files found in baseline folder"}

    return seed_baseline(files)


def seed_baseline(files: dict[str, bytes]) -> dict:
    """Parse baseline Excel files and insert rows into baseline_data table."""
    session = get_session()
    total_rows = 0
    snapshot_result = None

    try:
        # Clear previous baseline data
        session.query(BaselineData).delete()
        session.commit()

        for filename, content in files.items():
            # Skip non-Excel files (e.g. CLAUDE.md uploaded to baseline folder)
            if not filename.lower().endswith((".xlsx", ".xls")):
                continue

            data_type = _classify_file(filename)

            # Use targeted parsers for known file types
            if data_type == "outcomes":
                parsed_rows = _parse_partner_outcomes(content, filename)
            elif data_type == "provider_reconciliation":
                parsed_rows = _parse_provider_reconciliation(content, filename)
            else:
                parsed_rows = _parse_excel_to_rows(content, filename)

            for row in parsed_rows:
                record = BaselineData(
                    source_file=filename,
                    data_type=data_type,
                    sheet_name=row["sheet_name"],
                    row_data=row["row_data"],
                )
                session.add(record)
                total_rows += 1

        session.commit()

        # Always seed the budget snapshot with confirmed QB actuals
        snapshot_result = _seed_budget_snapshot()

        result = {
            "status": "success",
            "files_processed": len(files),
            "rows_inserted": total_rows,
        }
        if snapshot_result:
            result["budget_snapshot"] = snapshot_result
        return result
    except Exception as e:
        session.rollback()
        raise
    finally:
        session.close()
